# #=============================================================================================================================
# # Code for the Multi-commodity Flows problem (Arc version) for lectures of course AE4423 (Airline Planning & Optimisation)
# #=============================================================================================================================



from gurobipy import *
from openpyxl import *
from time import *


#=============================================================================================================================
#  Create the model which represent the nodes, arcs, and comodities 
#=============================================================================================================================
class Arc:
    def __init__(self,origin, destination, cost, capacity):
        self.From   = origin      #i
        self.To     = destination #j
        self.Cost   = cost
        self.Capac  = capacity

class Commodity:
    def __init__(self,origin, destination, quantity):
        self.From   = origin        #i
        self.To     = destination   #j
        self.Quant  = quantity

class Node:
    def __init__(self, id):
        self.id  = id
        self.InLinks  = [ ]         
        self.OutLinks = [ ]        
    
    def addInLink(self,Node):      # Add new 'In Link' to the node
        self.InLinks.append(Node)       
    
    def addOutLink(self,Node):     # Add new 'Out Link' to the node
        self.OutLinks.append(Node)



#=============================================================================================================================
#  Begin Auxiliar Functions
#=============================================================================================================================

#construct direct graph
def construct_graph():
      
    # Input excel file with arcs data (sheet1) and commodities data (sheet2)
    Arcs        = []
    Nodes       = []
    Commodities = []
    
    wb = load_workbook("Input_Lecture.xlsx", read_only=True)
    List_arcs = tuple(wb["Arcs"].iter_rows())
    List_commo = tuple(wb["Commodities"].iter_rows())
        
    # Generate Nodes
    NewNode = []
    for (ID,origin,destination,Cost,Capacity) in List_arcs[1:]:
        NewNode.append(int(origin.value))
        NewNode.append(int(destination.value))
    NewNode=set(NewNode)    #eliminate repeated nodes in the list
    for new in NewNode:
        Nodes.append(Node(new))

    # Insert arcs (and their characteristics) in a list of Arcs
    for (ID, origin, destination,Cost,Capacity) in List_arcs[1:]:
        Arcs.append(Arc (int(origin.value), int(destination.value), int(Cost.value), int(Capacity.value)))
        Nodes[int(origin.value)].addOutLink(int(destination.value))
        Nodes[int(destination.value)].addInLink(int(origin.value))

    # Insert commodities (and their characteristics) in a list of Commodities
    for (ID, origin, destination,quantity) in List_commo[1:]: 
        Commodities.append(Commodity (int(origin.value), int(destination.value), int(quantity.value)))

    return Arcs, Nodes, Commodities
    

#check if optimization was successful
def check_model_status(model):
    status = model.status
    if status != GRB.Status.OPTIMAL:
        if status == GRB.Status.UNBOUNDED:
            print('The model cannot be solved because it is unbounded')
        elif status == GRB.Status.INFEASIBLE:
            print('The model is infeasible; computing IIS')
            model.computeIIS()
            print('The following constraint(s) cannot be satisfied:')
            for c in model.getConstrs():
                if c.IISConstr:
                    print(c.constrName)
        elif status != GRB.Status.INF_OR_UNBD:
            print('Optimization was stopped with status',status)
        exit(0)


#print the result of the optimization into a readable text format
def print_model_result(x, model):
    for Arc in Arcs:
        Flow = 0
        for m in range(len(Commodities)):
            Flow += x[m,Arc.From,Arc.To].X
        if int(Flow)>0:
            print ('Arc(', Arc.From + 1, ',', Arc.To + 1, ')=', int(Flow))
    print ('Objective Function =', model.ObjVal/1.0)



#=============================================================================================================================
#  Optimization of the Model through Gurobi
#=============================================================================================================================
def MCF_Problem (Arcs, Nodes, Commodities):
    "Solve the MCF Problem with a Arc-based formulation using linear programming."
    
    # LP model (this is an object)
    model = Model("MCF")               
    
    # Decision Variables 
    # For addVar check https://www.gurobi.com/documentation/current/refman/py_model_addvar.html
    # vtype = 'C' for continous
    x = {}                    
    for Arc in Arcs:
        for k in range(len(Commodities)):
            x[k,Arc.From,Arc.To] = model.addVar(obj=Arc.Cost, vtype =GRB.CONTINUOUS, 
                                                name = ''.join(['Arc(', str(Arc.From), ',', str(Arc.To), ')']))

    # update model with the DVs before adding constraints
    model.update()                      


    # build 'balance' constraints
    # for .addConstr() check https://www.gurobi.com/documentation/current/refman/py_model_addconstr.html
    Balance = {}                      
    for k in range(len(Commodities)):
        for From in range(len(Nodes)):
            balance = quicksum(x[k,From,To] for To in Nodes[From].OutLinks) - quicksum(x[k,To,From] for To in Nodes[From].InLinks)
            if From == Commodities[k].From:
                Balance[k,From] = model.addConstr(balance == Commodities[k].Quant, 
                                                  name = ''.join(['Balance(',str(k), ',', str(From), ')']))
            elif From == Commodities[k].To:
                Balance[k,From] = model.addConstr(balance  == -Commodities[k].Quant, 
                                                  name = ''.join(['Balance(',str(k), ',', str(From), ')']))
            else:
                Balance[k,From] = model.addConstr(balance  == 0, 
                                                  name = ''.join(['Balance(',str(k), ',', str(From), ')']))


    # build 'capacity' constraints
    Capacity = {}                       
    for Arc in Arcs:
        Capacity[Arc.From,Arc.To] = model.addConstr(quicksum(x[k, Arc.From,Arc.To] for k in range(len(Commodities))) <= Arc.Capac, 
                                                   name = ''.join(['Capacity(', str(Arc.From), ',', str(Arc.To),')']))

    #save info to log file
    model.setParam("LogFile", 'log_file')
    #update gurobi with the constraints
    model.update()
    # Useful for model debugging
    model.write("MCF_Model.lp")
    model.optimize()
    
    check_model_status(model)
    print_model_result(x, model)
    
    


#=============================================================================================================================
#  Run the Program
#=============================================================================================================================
if __name__ == '__main__':
    Arcs, Nodes, Commodities = construct_graph()
   
    # RUN MCF PROBLEM
    start_time = time()
    MCF_Problem(Arcs, Nodes, Commodities)    
    print ('Run Time =', time() - start_time)






