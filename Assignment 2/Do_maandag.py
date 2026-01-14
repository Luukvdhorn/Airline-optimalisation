import numpy as np
import pandas as pd
import openpyxl
import math
import matplotlib.pyplot as plt 


# ================ Data Import =================
#Dataset demand data, airport characteristics   
wb = openpyxl.load_workbook("Assignment 2/DemandGroup40.xlsx", data_only=True)
ws = wb.active

icao_row = 5    
lat_row = 6    
lon_row = 7    
start_col = 3
runway_row = 8
runways = []
hub_index = 2

airports = []
latitudes = []
longitudes = []

#Latitude, Longitude and Runway length import
col = start_col
while True:
    icao = ws.cell(row=icao_row, column=col).value
    if icao is None:
        break
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    runway = ws.cell(row=runway_row, column=col).value

    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    runways.append(float(runway))
    col += 1

runway_arrival = np.array(runways)
n = len(airports)


#Distance formula to construct distance matrix
def distance(phi_i, lam_i, phi_j, lam_j):
    RE = 6371.0
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 + np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))

d = np.zeros((n, n))
for i in range(n):
    for j in range(n):
        d[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])


# Demand matrix import
demand_start_row = icao_row + 7
demand_start_col = start_col

D = np.zeros((n, n))
for i in range(n):
    for j in range(n):
        val = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(val) if val is not None else 0
        except:
            D[i,j] = 0

# Dataset hour coefficients import
wb_h = openpyxl.load_workbook("Assignment 2/HourCoefficients.xlsx", data_only=True)
ws_h = wb_h.active

T = 24  # 24 hours
H = np.zeros((n, T))
start_row = 3
start_col = 4
for i in range(n):
    for t in range(T):
        val = ws_h.cell(row=start_row + i, column=start_col + t).value
        try:
            H[i, t] = float(val) if val is not None else 0
        except:
            H[i, t] = 0

# Calculate demand per hour by spreading daily demand using hour coefficients
dem_hour = np.zeros((n, n, T))   # demand per origin-dest-hour
for i in range(n):
    for j in range(n):
        for t in range(T):
            dem_hour[i, j, t] = D[i, j] * H[i, t]


#Dataset Fleet data import
wb2 = openpyxl.load_workbook("Assignment 2/FleetType.xlsx", data_only=True)
sheet2 = wb2.active

aircraft_names = [cell.value for cell in sheet2[1][1:] if cell.value is not None]
n_ac = len(aircraft_names)
data = {}
for row in sheet2.iter_rows(min_row=2, values_only=True):
    pname = row[0]
    if pname is None:
        continue
    values = row[1:1 + n_ac]
    data[pname] = values

df_aircraft = pd.DataFrame(data, index=aircraft_names)


# Extract parameters from aircraft data 
s = df_aircraft['Seats'].values
v = df_aircraft['Speed [km/h]'].values
ra = df_aircraft['Maximum Range [km]'].values
RAC = df_aircraft['Runway Required [m]'].values
TAT = df_aircraft['Average TAT [min]'].values
fleet = df_aircraft['Fleet'].values.astype(int)
cl = df_aircraft['Lease Cost [€/day]'].values
C_fix = df_aircraft['Fixed Operating Cost (Per Fligth Leg)  [€]'].values
CT = df_aircraft['Cost per Hour'].values
CF = df_aircraft['Fuel Cost Parameter'].values



# ================ Parameters and Functions =================

step_minutes = 6        #time step in minutes
total_steps = 24 * 10   # 6 min steps in 24h

#Function to convert timestep to hour
def timestep_to_hour(timestep):
    total_minutes = timestep * step_minutes
    hour = total_minutes // 60
    return hour

#Function to check if action fulfils constraints (runway length and range)
def action_possible(stage, ac_type):
    possible = []
    if stage == hub_index:
        candidates = range(n)
    else:
        candidates = [stage, hub_index]
    for dest in candidates:
        if RAC[ac_type] < runway_arrival[dest] and ra[ac_type] >= d[stage, dest]:
            possible.append(dest)
    return possible

# Function to calculate block time [min] of one flight
def block_time(origin, dest, ac_type):
    if origin == dest:
        return 0
    bt = 15 + (d[origin,dest] / v[ac_type]) * 60 + TAT[ac_type] + 15
    return bt

# Function to calculate operating cost of one flight 
def operating_cost(origin, dest, ac_type):
    if origin == dest:
        return 0
    C_fixed = C_fix[ac_type]
    C_time = CT[ac_type] * (d[origin,dest] / v[ac_type])
    C_fuel = (CF[ac_type] * 1.42 / 1.5) * d[origin,dest]
    tot_cost = C_fixed + C_time + C_fuel
    return tot_cost

# Function to calculate revenue of one flight
def revenue_func(origin, dest, flow):
    if origin == dest:
        return 0
    y = 5.9 * d[origin,dest]**(-0.76) + 0.043
    rev = y * d[origin,dest] * flow
    return rev

# Function to calculate captured demand for a flight (including previous hours)
def capture_demand(origin, dest, timestep, ac_type, dem_hour):
    hour, _ = divmod(timestep * step_minutes, 60)
    capacity = 0.8 * s[ac_type]
    remaining_capacity = capacity
    flow = 0
    for h in [hour, hour-1, hour-2]:
        if h < 0:
            continue
        available = dem_hour[origin, dest, h]
        if available <= 0:
            continue
        taken = min(available, remaining_capacity)
        taken_int = math.floor(taken)  
        if taken_int == 0:
            continue 
        dem_hour[origin, dest, h] -= taken_int
        flow += taken_int
        remaining_capacity -= taken_int
        if remaining_capacity <= 0:
            break
    return flow, dem_hour



# ================ Dynamic Programming and Scheduling =================

# Function for dynamic programming algorithm
def dynamic_programming(ac_type, dem_hour):
    profit_matrix = np.full((n, total_steps), -1e9)   # Storage space for profit
    action_matrix = np.full((n, total_steps), -1)     # Storage of action taken

    # Last run only profitable if it goes to the hub
    profit_matrix[:, -1] = -1e7
    profit_matrix[hub_index, -1] = 0

    for t in range(total_steps-2, -1, -1):
        current_hour = timestep_to_hour(t)
        for loc in range(n):
            best_profit = -1e6
            best_action = -1

            if t >= total_steps - 1:
                if loc == hub_index:
                    candidate_dests = [hub_index]
                else:
                    candidate_dests = [hub_index]
            else:
                candidate_dests = action_possible(loc, ac_type)

            for dest in candidate_dests:
                if dest == loc:
                    profit = profit_matrix[loc, t+1]
                    flow = 0
                else:
                    blockt = block_time(loc, dest, ac_type)
                    arrival_time = t + math.ceil(blockt/step_minutes)
                    if arrival_time >= total_steps:
                        future_profit = -1e9
                    else:
                        future_profit = profit_matrix[dest, arrival_time]

                    expected_demand = 0
                    for hh in [current_hour, current_hour-1, current_hour-2]:
                        if hh >= 0:
                            expected_demand += dem_hour[loc,dest,hh]
                    capacity = 0.8 * s[ac_type]
                    flow = min(capacity, expected_demand)
                    revenue = revenue_func(loc, dest, flow)
                    cost = operating_cost(loc, dest, ac_type)
                    profit = revenue - cost + future_profit

                if profit > best_profit:
                    best_profit = profit
                    best_action = dest

            profit_matrix[loc, t] = best_profit
            action_matrix[loc, t] = best_action

    return action_matrix, profit_matrix

# Function to schedule flights based on action and profit matrices
def schedule(ac_type, action_matrix, profit_matrix, dem_hour):
    route = [(0, hub_index)]
    current_pos = hub_index
    t = 0
    total_block_time = 0
    flows = []

    while t < total_steps:
        next_pos = int(action_matrix[current_pos, t])
        if next_pos < 0 or next_pos >= n:
            break
        if next_pos == current_pos:            
            t += 1
            continue
        bt = block_time(current_pos, next_pos, ac_type)
        t_next = t + math.ceil(bt/step_minutes)
        if t_next >= total_steps:
            break
        total_block_time += bt
        flow, dem_hour = capture_demand(current_pos, next_pos, t, ac_type, dem_hour)
        
        revenue_leg = revenue_func(current_pos, next_pos, flow)
        cost_leg = operating_cost(current_pos, next_pos, ac_type)
        profit_leg = revenue_leg - cost_leg

        print(f"[SCHEDULE] AC {ac_type}: Vlucht {airports[current_pos]}->{airports[next_pos]} flow={flow:.1f} "
            f"revenue={revenue_leg:.2f} cost={cost_leg:.2f} profit_leg={profit_leg:.2f}")
        
        flows.append(flow)
        route.append((t_next, next_pos))
        current_pos = next_pos
        t = t_next

    min_block = 6*60
    if total_block_time < min_block:
        print(f"[SCHEDULE] Aircraft type {ac_type} totale blocktime {total_block_time:.1f} min < minimale blocktijd {min_block} min, winst wordt -1e9 gezet")
        profit = -1e9
    else:
        profit = profit_matrix[hub_index, 0] - cl[ac_type]

    return route, profit, total_block_time, flows, dem_hour






# ================ Main Loop to build optimal schedule ==================

# Main loop to build the flight schedule
demand = dem_hour.copy()
available_ac = fleet.copy()
total_passengers_transported = 0
solution_dict = {}
iteration = 0

used_ac_count = np.zeros(len(fleet), dtype=int)
total_profit = 0  # Voor cumulatieve winst

while any(available_ac > 0):
    profits = np.full(n_ac, -1e8)
    routes = {}

    for k in range(n_ac):
        if available_ac[k] > 0:
            action_mat, profit_mat = dynamic_programming(k, demand)
            r, p, t_block, flown, d_new = schedule(k, action_mat, profit_mat, demand)
            routes[k] = (r, p, t_block, flown, d_new)
            profits[k] = p

    # Print status per iteration
    print(f"Iteration: {iteration} - Profits: {profits}, Available AC: {available_ac}")

    if np.all(profits < 0):
        print("No profitable flights left, stop.")
        break

    k_best = np.argmax(profits)
    r, p, t_block, flown, d_new = routes[k_best]
    demand = d_new 
    available_ac[k_best] -= 1
    used_ac_count[k_best] += 1
    total_profit += p
    total_passengers_transported += sum(flown)
    solution_dict[f"Route {iteration}"] = {"Aircraft type": k_best, "Profit": p,
                                        "Utilization": t_block,
                                        "Route": r,
                                        "Flows": flown}

    print(f"[MAIN LOOP] Iteration {iteration}: Added plane type {k_best} with profit {p:.1f}, block time {t_block:.1f} min")
    print(f"Remaining demand sum: {np.sum(demand):.1f}")
    print(f"Total passengers transported so far: {total_passengers_transported:.1f}")

    iteration += 1

print(f"\nTotal profit for all flights: {total_profit:.2f} euro")





# ================ Print results and Analysis ==================

print("\n===== Used aircraft per type =====")
for idx, ac_name in enumerate(df_aircraft.index):
    print(f"{ac_name}: {used_ac_count[idx]} used of {fleet[idx]} available")

# Function to convert timestep to label for printing
def timestep_to_label(ts, timestep_duration=6):
    minutes = ts * timestep_duration
    h = minutes // 60
    m = minutes % 60
    return f"{int(h):02d}:{int(m):02d}"

def print_all_routes(solution_dict, airports, timestep_to_label):
    print("\n=== Routes for optimal schedule ===")

    for route_name, sched in solution_dict.items():
        route = sched["Route"]                
        ac_type = sched["Aircraft type"]
        flown = sched.get("Flows", [])

        print(f"\n{route_name} - Aircraft type {ac_type}:")

        prev_arr_ts = None 

        for i in range(len(route) - 1):
            dep_ts, origin = route[i]
            arr_ts, dest   = route[i + 1]

            passengers = flown[i] if i < len(flown) else 0

            block_t = block_time(origin, dest, ac_type)

            if (arr_ts - dep_ts) * 6 > block_t:
                dep_ts = arr_ts - block_t / 6

            if prev_arr_ts is not None and dep_ts != prev_arr_ts:
                waiting_min = (dep_ts - prev_arr_ts) * 6
                print(f"    Waiting time: {math.floor(waiting_min):.0f} min")

            print(
                f"  Departure {timestep_to_label(dep_ts)} from {airports[origin]} "
                f"→ arrival {timestep_to_label(arr_ts)} at {airports[dest]} | "
                f"Block time: {math.ceil(block_t):.0f} min | "
                f"Passengers: {passengers:.0f}"
            )

            prev_arr_ts = arr_ts


print_all_routes(solution_dict, airports, timestep_to_label)



print("\n=========KPI's flight schedule:=========")
ASK = 0.0
RPK = 0.0

TOTAL_REVENUE = 0.0
TOTAL_COST = 0.0
TOTAL_LEASE = 0.0

for sched in solution_dict.values():
    route = sched["Route"]
    flows = sched["Flows"]
    ac_type = sched["Aircraft type"]
    seats = s[ac_type]

    TOTAL_LEASE += cl[ac_type]   

    for i, passengers in enumerate(flows):
        if passengers <= 0:
            continue

        _, origin = route[i]
        _, dest   = route[i+1]

        distance_leg = d[origin, dest]

        ASK += seats * distance_leg
        RPK += passengers * distance_leg

        y = 5.9 * distance_leg**(-0.76) + 0.043
        revenue_leg = y * distance_leg * passengers
        TOTAL_REVENUE += revenue_leg

        cost_leg = operating_cost(origin, dest, ac_type)
        TOTAL_COST += cost_leg

TOTAL_COST += TOTAL_LEASE
CASK = TOTAL_COST / ASK 
RASK = TOTAL_REVENUE / ASK 
YIELD = TOTAL_REVENUE / RPK
ANLF = RPK / ASK 
BELF = CASK / YIELD 

print(f"ASK   : {ASK:,.0f} seat-km")
print(f"RPK   : {RPK:,.0f} pax-km")
print(f"CASK  : {CASK:.4f} €/ASK")
print(f"RASK  : {RASK:.4f} €/ASK")
print(f"YIELD : {YIELD:.4f} €/RPK")
print(f"ANLF  : {ANLF:.3f}")
print(f"BELF  : {BELF:.3f}")

print(f"\nTotal revenue: {TOTAL_REVENUE:.2f}")
print(f"Total costs: {TOTAL_COST:.2f}")
print(f"Total profit: {TOTAL_REVENUE-TOTAL_COST:.2f}")

print("\n=== Number of flights between two airports ===")

flight_counts = {}

for sched in solution_dict.values():
    route = sched["Route"]

    for i in range(len(route) - 1):
        origin = route[i][1]
        dest   = route[i + 1][1]

        key = (airports[origin], airports[dest])
        flight_counts[key] = flight_counts.get(key, 0) + 1

for (origin, dest), count in flight_counts.items():
    print(f"{origin} → {dest}: {count} flights")

print("\n=== Distances between key airports ===")
print(f' Amsterdam - London {d[2,0]} km')
print(f' Amsterdam - Paris {d[2,1]} km')
print(f' Amsterdam - Munich {d[2,6]} km')
print(f' Amsterdam - Berlin {d[2, 11]} km')
print(f' Amsterdam - Frankfurt {d[2,3]} km')
print(f' Amsterdam - Dublin {d[2, 8]} km')




# ================ Visualization of Schedule ==================


fig, ax = plt.subplots(figsize=(16, 6))

ax.set_yticks(range(len(airports)))
ax.set_yticklabels(airports)
ax.set_ylim(-1, len(airports))

timestep_duration = 6
total_timesteps = 24 * 60 // timestep_duration

ax.set_xlim(0, total_timesteps)

minor_ticks = range(0, total_timesteps + 1, 2 * 60 // timestep_duration)
ax.set_xticks(minor_ticks)
ax.set_xticklabels([timestep_to_label(t) for t in minor_ticks], rotation=45)
ax.tick_params(axis='x', labelsize=8)

colors = ['blue', 'orange', 'green', 'red', 'purple', 'brown', 'pink', 'gray']

for idx, (route_name, sched) in enumerate(solution_dict.items()):
    route = sched["Route"]
    ac_type = sched["Aircraft type"]
    color = colors[idx % len(colors)]

    prev_arr_ts = None
    prev_dest = None

    for i in range(len(route) - 1):
        dep_ts, origin = route[i]
        arr_ts, dest   = route[i + 1]

        block_t = block_time(origin, dest, ac_type)

        if (arr_ts - dep_ts) * 6 > block_t:
            dep_ts = arr_ts - block_t / 6

        if prev_arr_ts is not None and dep_ts > prev_arr_ts:
            ax.plot(
                [prev_arr_ts, dep_ts],
                [origin, origin],
                color=color,
                linewidth=2
            )

        ax.plot(
            [dep_ts, arr_ts],
            [origin, dest],
            color=color,
            linewidth=2,
            label=f"{route_name} - AC type {ac_type}" if i == 0 else ""
        )

        prev_arr_ts = arr_ts
        prev_dest = dest

ax.set_xlabel("Time")
ax.set_ylabel("Airports")
ax.set_title("Aircraft Timetable (Linear Flight & Waiting Times)")
ax.grid(True)
ax.legend()
plt.tight_layout()
plt.show()
