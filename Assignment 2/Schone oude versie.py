import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import * 
import openpyxl
import math
import copy

wb = load_workbook("Assignment 2/DemandGroup40.xlsx", data_only=True)
ws = wb.active

icao_row = 5    
lat_row = 6    
lon_row = 7    
start_col = 3
runway_row = 8
runways = []
hub_index = 2

RE = 6371.0                                     # Radius Earth
f = 1.42                                        # Fuel cost constant 

airports = []
latitudes = []
longitudes = []


col = start_col

while True:
    icao = ws.cell(row=icao_row, column=col).value
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    runway = ws.cell(row=runway_row, column=col).value
    if icao is None:
        break
    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    runways.append(float(runway))
    col += 1

#DISTANCE FORMULA

def distance(phi_i, lam_i, phi_j, lam_j):
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 +np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))


n = len(airports)
dij = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        dij[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])

#DEMAND IMPORT
demand_start_row = icao_row + 7
demand_start_col = start_col 

n = len(airports)
D = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        cell = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(cell) if cell is not None else 0
        except:
            D[i, j] = 0


#FLEET IMPORT. --> Heel anders gedaan dan inladen van andere data: aanpassen zodat het een geheel wordt
wb2 = openpyxl.load_workbook("Assignment 2/FleetType.xlsx", data_only=True)
sheet2 = wb2.active

aircraft_names = [cell.value for cell in sheet2[1][1:]if cell.value is not None]
n_ac = len(aircraft_names)
data = {}

for row in sheet2.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue  
    parameter_name = row[0]
    values = row[1:1 + n_ac]  
    data[parameter_name] = values

df_aircraft = pd.DataFrame(data, index=aircraft_names) #vooral hier anders met aanmaken nieuwe dataset

#HOUR COEFFICIENT
wb = openpyxl.load_workbook("Assignment 2/HourCoefficients.xlsx", data_only=True)
ws = wb.active

start_row = 3        #AMSTERDAM
start_col = 4        #Hour 0 
n = len(airports)    #LENGTE RIJEN
T = 24               #AANTAL UREN (D t/m AA)

H = np.zeros((n, T))

for i in range(n):
    for t in range(T):
        cell = ws.cell(row=start_row + i, column=start_col + t).value
        try:
            H[i, t] = float(cell) if cell is not None else 0
        except:
            H[i, t] = 0


# DATA IMPORT
N = range(len(airports))                    # Set of airports; i, j in N
K = range(len(df_aircraft))                 # Set of aircrafts; k in K
ac = len(df_aircraft)                       # Total amount of aircrafts


d = np.zeros((n, n))                        # Distance between i and j
for i in N:
    for j in N:
        d[i, j] = dij[i, j]


y = np.zeros((n, n))
for i in N:                                 #YIELD
    for j in N:
        if d[i, j] > 0:
            y[i, j] = (5.9 * d[i, j]**(-0.76) + 0.043)
        else:
            y[i, j] = 0 

v = np.zeros(ac)                            #SPEED OF AIRPLANES (k)
for k in K:
    v[k] = df_aircraft['Speed [km/h]'].iloc[k]
    
s = np.zeros(ac)                            #SEATS
for k in K:
    s[k] = df_aircraft['Seats'].iloc[k]

ra = np.zeros(ac)                           #RANGE OF AIRCRAFT
for k in K:
    ra[k] = df_aircraft['Maximum Range [km]'].iloc[k]

RAC = np.zeros(ac)                          #RUNWAY REQUIREMENT DEPARTURE
for k in K:
    RAC[k] = df_aircraft['Runway Required [m]'].iloc[k]

RAP = np.zeros(n)                           #RUNWAY REQUIREMENT ARRIVAL
for j in N:
    RAP[j] = runways[j]

TAT = np.zeros(ac)                                                      #TURN AROUND TIMES IN MIN
for k in K:
    TAT[k] = df_aircraft['Average TAT [min]'].iloc[k]

Fleet = np.zeros(ac)                                                      #Fleettypes
for k in K:
    Fleet[k] = df_aircraft['Fleet'].iloc[k]

#COSTS 

cl = np.zeros(ac)                           #COSTS WEEKLY LEASE
for k in K:
    cl[k] = df_aircraft['Lease Cost [€/day]'].iloc[k]
    

# C_fix = np.zeros(ac)                            #COST FIXED OPERATING
# for k in K:
#     C_fix[k] = df_aircraft['Fixed Operating Cost (Per Fligth Leg)  [€]'].iloc[k]
    

# CT = np.zeros(ac)                           #TIME COST PARAMETER
# for k in K:
#     CT[k] = df_aircraft['Cost per Hour'].iloc[k]

# CF = np.zeros(ac)                           #FUEL COST PARAMETER
# for k in K:
#     CF[k] = df_aircraft['Fuel Cost Parameter'].iloc[k]   

C_fix = df_aircraft['Fixed Operating Cost (Per Fligth Leg)  [€]'].values
CT    = df_aircraft['Cost per Hour'].values
CF    = df_aircraft['Fuel Cost Parameter'].values
v     = df_aircraft['Speed [km/h]'].values



a = {}                                  #BIG M CONSTRAIN FOR RUNWAYS
for i in N:
    for j in N:
        for k in K:
            if d[i,j] <= ra[k]:
                a[i,j,k] = 10000
            else:
                a[i,j,k] = 0

# g = np.ones(n)                      #HUB IS AMSTERDAM 
# g[hub_index] = 0 #hebben we dit nog nodig?

LF = 0.80               # van 0.75 naar 0.80


# constraints airports functie (runway, range, minimaal 6 uur)
    #stage = index van luchthaven
    #required runway for aircraft < availible runway AND range arcraft 
def  action_possible(stage, aircraft_type):  
    possible_destinations = []

    #optie 1: aircraft is at hub 
    if stage == hub_index:
        candidate_destinations = range(len(airports)) # (choose other airport (can be hub))
    #optie 2: aircraft is not at hub 
    else:
        candidate_destinations = [stage, hub_index] # (stay or back to hub)


    for destination in candidate_destinations:
            if (RAC[aircraft_type] < RAP[destination] and ra[aircraft_type] >= d[stage, destination]):
                possible_destinations.append(destination)
    return(possible_destinations)

#TEST als je vanaf hub kijkt heb je met vliegtuig 1 20 opties
print(f'test action: {action_possible(2, 1)}')

#Time calculation --> function to calculate blocking time of one fligth --> sum minimal 6 hours after assigning routes Denk ik (ros)
def block_time(airport_from, airport_to, aircraft_type):
    if airport_from == airport_to: #als vliegtuig blijft staan, geen blocking time
        BT = 0
    else:
        BT = 15 + (d[airport_from, airport_to] / v[aircraft_type]) * 60 + TAT[aircraft_type] + 15
    return BT

#TEST berekenen blocking time: van ams naar ams met aircraft 1 
print(f'test bt: {block_time(2, 4, 1):.2f} min, {block_time(2, 4, 1) / 60:.2f} uur')

#Scheduling horizon: time steps 
#time step has to be converted to a time 00:06, 00:12 etc
def timestep_converting(timestep):
    step = 6 # 6 min per step
    steps_day = 24 * 10
    total_minutes = (timestep % steps_day) * step # tijdstap binnen de dag
    
    hours = total_minutes // 60
    minutes = total_minutes % 60
    time = f"{hours:02d}:{minutes:02d}"
    return hours, minutes 

#TEST convert timestep to tijdstip voor timetable: LET OP: step 1 = 00:00 
print(f'test time= {timestep_converting(3)}')


#TO do:

#denk dat het handig is om de berekeningen die we eerder doen wel om te schrijven naar functies, dat is makkelijker met aanroepen

#demand function

#cost function
def operating_costs(orgin, destination, aircraft_type):
    
    C_t = CT[aircraft_type]
    C_f = CF[aircraft_type]
    Speed = v[aircraft_type]

    C_fixed = C_fix[aircraft_type]
    if orgin == destination:                                    # Heb even toegevoegd dat als het vliegtuig niet tussen airports gaat fixed cost ook 0 zijn, want staat stil
        C_fixed = 0
    C_time = C_t * d[orgin, destination] / Speed
    C_fuel = ((C_f * 1.42 / 1.5) * d[orgin, destination])

    calculate_operating_cost = C_fixed + C_time + C_fuel

    return calculate_operating_cost

print(f'Operation cost €{operating_costs(2, 7, 1):.2f}')
            
#TEST OM ALLE OPERATING COST TE PRINTEN (als de vliegtuigen groter worden wordt ook cost hoger ;), ook gecheckt met oude manier, zijn gelijk.
#k = 1 
#df_cost = pd.DataFrame(Ck_ij[:, :, k], index=airports, columns=airports)
#print(df_cost.round(0))

#revenue function
def revenue_function(orgin, destination, flow):
    revenue = (5.9 * d[orgin, destination]**(-0.76) + 0.043) * flow
    return revenue

#Hier moet later wellicht nog de passagiers per time step bij? Maar dat weet ik nu niet zo goed. 
# Heb flow toegevoegd, dit fikst dat. In Dynamic programming word flow als goed is bepaald.

print(f'Revenue €{revenue_function(2, 3, 10):.2f}')

def initial_demand(orgin, destination, hour):     # DIT WAS ZO VAN DIT KRIJG JE UIT GEWOON INEZEN
    total_demand = D[orgin, destination]
    demand = H[orgin, hour] * total_demand
    return demand

print(f'Demand is {initial_demand(1, 2, 5)}')

def usable_demand(orgin, destination, hour):       # HIER VOEG JE TOE DAT JE UIT DE UREN ER VOOR OOK KAN KRIJGEN. PROBLEEM IS ALLEEN DAT IK HEIR NOG GEEN REKENING HIELD MET 
    start_demand = initial_demand(orgin, destination, hour)   # DAT JE EERST KIJK NAAR T ALS ER DAN NOG PLEK IS NAAR T -1 EN ALS ER DAN NOG PLEK IS NAAR T -2
    use_demand = start_demand + initial_demand(orgin, destination, hour-1) + initial_demand(orgin, destination, hour-2)
    return use_demand

print(f'Demand with 2 previous time slots is {usable_demand(1, 2, 5)}')


def dynamic_programming(aircraft_type):
    total_steps = 24 * 10           # Amount of steps in the day
    end_time = 24 * 60              # Time corresponding to last step

    profit_matrix = np.zeros((len(airports), total_steps))      # Storage space for profit
    action_matrix = np.zeros((len(airports), total_steps))      # Storage of action taken

    for i in range(total_steps):
        current_time_step = total_steps - (i+1)                 # Starts at the end, but the very last is already determined
        current_time = current_time_step * 6

        if current_time_step % 60 == 0:
            print(current_time_step)                            # Control for if the function is going trough all time steps
        
        for j in range(len(airports)):
            best_action = -1                                    # Set base for action value, if action is  chosen it will get a positive value

            if j == hub_index and i == 0:
                best_solution = 0                               # For start position best solution is 0, is given
            else:
                best_solution = -10e5                           # Base every where big negative so it can only improve

            for k in action_possible(j, aircraft_type):

                if k == j:
                    if current_time_step < total_steps - 1:
                        current_solution = profit_matrix[j][current_time_step + 1]      # Previous profit is same as current
                    else:
                        current_solution = -10e5

                else:
                     fligth_duration = block_time(j, k, aircraft_type)
                     arrival_time = current_time + fligth_duration
                     arrival_time_step = math.ceil(arrival_time / 6) - 1
                     if arrival_time_step >= total_steps:
                        future_profit = 0
                     else:
                        future_profit = profit_matrix[k][arrival_time_step]
                     
                     opertion_cost = operating_costs(j, k, aircraft_type)
                     current_hour, current_min = timestep_converting(current_time_step)
                     demand = usable_demand(j, k, current_hour)
                     revenue = revenue_function(j, k, min(demand, s[aircraft_type]))
                     profit = revenue - opertion_cost
                     arrival_time_step = math.ceil(arrival_time/6)-1
                     current_solution = profit + future_profit

                if current_solution > best_solution:
                    best_solution = current_solution
                    best_action = k

                action_matrix[j][current_time_step] = best_action
                profit_matrix[j][current_time_step] = best_solution

    return action_matrix, profit_matrix


print(f'Poging {dynamic_programming(2)}')

def scheduel(aircraft_type, action_matrix, profit_matrix):
    optimal_route = [[0,2]]                 # start time step (0) and Start airport (2 = hub)
    time_step = 0
    current_airport = 2
    operation_time = 0

    while time_step <= 24 * 10:              # 24 * 10 is the last time step
        next_airport = int(action_matrix[current_airport][time_step])
        if next_airport == current_airport:
            next_time_step = time_step + 1
        else:
            next_time_step = time_step + math.floor(block_time(current_airport, next_airport, aircraft_type))
            operation_time += block_time(current_airport, next_airport, aircraft_type)
            optimal_route.append([next_time_step, next_airport])

        current_airport = int(next_airport)
        time_step = next_time_step

    profit = profit_matrix[1][0] - cl[aircraft_type]

    return optimal_route, profit, operation_time


if __name__ == "__main__":
    # Kies een aircraft‐type (bv. 0, 1 of 2)
    ac_type = 2

    # 1) Voer DP uit
    action_matrix, profit_matrix = dynamic_programming(ac_type)

    # 2) Simuleer je optimale route
    route, profit, block = scheduel(ac_type, action_matrix, profit_matrix)

    print("Optimale route (ts, loc):", route)
    print(f"Verwachte profit = €{profit:.2f}")
    print(f"Totale block‐tijd = {block:.1f} min")