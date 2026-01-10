import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import * 
import openpyxl
import math
import copy

wb = load_workbook("Assignment 2/DemandGroup40.xlsx", data_only=True)
ws = wb.active

icao_row = 5    
lat_row = 6    
lon_row = 7    
start_col = 3
runway_row = 8
runways = []
hub_index = 2

RE = 6371.0                                     # Radius Earth
f = 1.42                                        # Fuel cost constant 

airports = []
latitudes = []
longitudes = []


col = start_col

while True:
    icao = ws.cell(row=icao_row, column=col).value
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    runway = ws.cell(row=runway_row, column=col).value
    if icao is None:
        break
    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    runways.append(float(runway))
    col += 1

#DISTANCE FORMULA

def distance(phi_i, lam_i, phi_j, lam_j):
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 +np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))


n = len(airports)
dij = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        dij[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])
print(dij[2,3])

#DEMAND IMPORT
demand_start_row = icao_row + 7
demand_start_col = start_col 

n = len(airports)
D = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        cell = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(cell) if cell is not None else 0
        except:
            D[i, j] = 0


#FLEET IMPORT. --> Heel anders gedaan dan inladen van andere data: aanpassen zodat het een geheel wordt
wb2 = openpyxl.load_workbook("Assignment 2/FleetType.xlsx", data_only=True)
sheet2 = wb2.active

aircraft_names = [cell.value for cell in sheet2[1][1:]if cell.value is not None]
n_ac = len(aircraft_names)
data = {}

for row in sheet2.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue  
    parameter_name = row[0]
    values = row[1:1 + n_ac]  
    data[parameter_name] = values

df_aircraft = pd.DataFrame(data, index=aircraft_names) #vooral hier anders met aanmaken nieuwe dataset

print(df_aircraft)

#HOUR COEFFICIENT
wb = openpyxl.load_workbook("Assignment 2/HourCoefficients.xlsx", data_only=True)
ws = wb.active

start_row = 3        #AMSTERDAM
start_col = 4        #Hour 0 
n = len(airports)    #LENGTE RIJEN
T = 24               #AANTAL UREN (D t/m AA)

H = np.zeros((n, T))

for i in range(n):
    for t in range(T):
        cell = ws.cell(row=start_row + i, column=start_col + t).value
        try:
            H[i, t] = float(cell) if cell is not None else 0
        except:
            H[i, t] = 0

# DATA IMPORT
N = range(len(airports))                    # Set of airports; i, j in N
K = range(len(df_aircraft))                 # Set of aircrafts; k in K
ac = len(df_aircraft)                       # Total amount of aircrafts


d = np.zeros((n, n))                        # Distance between i and j
for i in N:
    for j in N:
        d[i, j] = dij[i, j]


y = np.zeros((n, n))
for i in N:                                 #YIELD
    for j in N:
        if d[i, j] > 0:
            y[i, j] = (5.9 * d[i, j]**(-0.76) + 0.043)
        else:
            y[i, j] = 0 

v = np.zeros(ac)                            #SPEED OF AIRPLANES (k)
for k in K:
    v[k] = df_aircraft['Speed [km/h]'].iloc[k]
    
s = np.zeros(ac)                            #SEATS
for k in K:
    s[k] = df_aircraft['Seats'].iloc[k]

ra = np.zeros(ac)                           #RANGE OF AIRCRAFT
for k in K:
    ra[k] = df_aircraft['Maximum Range [km]'].iloc[k]

RAC = np.zeros(ac)                          #RUNWAY REQUIREMENT DEPARTURE
for k in K:
    RAC[k] = df_aircraft['Runway Required [m]'].iloc[k]

RAP = np.zeros(n)                           #RUNWAY REQUIREMENT ARRIVAL
for j in N:
    RAP[j] = runways[j]

TAT = np.zeros(ac)                                                      #TURN AROUND TIMES IN MIN
for k in K:
    TAT[k] = df_aircraft['Average TAT [min]'].iloc[k]

Fleet = np.zeros(ac)                                                      #Fleettypes
for k in K:
    Fleet[k] = df_aircraft['Fleet'].iloc[k]

#COSTS 

cl = np.zeros(ac)                           #COSTS WEEKLY LEASE
for k in K:
    cl[k] = df_aircraft['Lease Cost [€/day]'].iloc[k]
    

C_fix = df_aircraft['Fixed Operating Cost (Per Fligth Leg)  [€]'].values
CT    = df_aircraft['Cost per Hour'].values
CF    = df_aircraft['Fuel Cost Parameter'].values
v     = df_aircraft['Speed [km/h]'].values



a = {}                                  #BIG M CONSTRAIN FOR RUNWAYS
for i in N:
    for j in N:
        for k in K:
            if d[i,j] <= ra[k]:
                a[i,j,k] = 10000
            else:
                a[i,j,k] = 0

LF = 0.80               # van 0.75 naar 0.80


#VANAF HIER CODE DYNAMIC PROGRAMMEREN 
 
def  action_possible(stage, aircraft_type):  
    possible_destinations = []

    #optie 1: aircraft is at hub 
    if stage == hub_index:
        candidate_destinations = range(len(airports)) # (choose other airport (can be hub))
    #optie 2: aircraft is not at hub 
    else:
        candidate_destinations = [stage, hub_index] # (stay or back to hub)

    for destination in candidate_destinations:
            if (RAC[aircraft_type] < RAP[destination] and ra[aircraft_type] >= d[stage, destination]):
                possible_destinations.append(destination)
    return(possible_destinations)

#TEST als je vanaf hub kijkt heb je met vliegtuig 1 20 opties
print(f'test action: {action_possible(2, 1)}')

#Time calculation --> function to calculate blocking time of one fligth --> sum minimal 6 hours after assigning routes Denk ik (ros)
def block_time(airport_from, airport_to, aircraft_type):
    if airport_from == airport_to: #als vliegtuig blijft staan, geen blocking time
        BT = 0
    else:
        BT = 15 + (d[airport_from, airport_to] / v[aircraft_type]) * 60 + TAT[aircraft_type] + 15
    return BT

#TEST berekenen blocking time: van ams naar ams met aircraft 1 
print(f'test bt: {block_time(2, 4, 1):.2f} min, {block_time(2, 4, 1) / 60:.2f} uur')

#Scheduling horizon: time steps 
#time step has to be converted to a time 00:06, 00:12 etc
def timestep_converting(timestep):
    step = 6 # 6 min per step
    steps_day = 24 * 10
    total_minutes = (timestep % steps_day) * step # tijdstap binnen de dag
    
    hours = total_minutes // 60
    minutes = total_minutes % 60
    time = f"{hours:02d}:{minutes:02d}"
    return hours, minutes 

#TEST convert timestep to tijdstip voor timetable: LET OP: step 1 = 00:00 
print(f'test time= {timestep_converting(3)}')

def operating_costs(orgin, destination, aircraft_type):
    
    C_t = CT[aircraft_type]
    C_f = CF[aircraft_type]
    Speed = v[aircraft_type]

    C_fixed = C_fix[aircraft_type]
    if orgin == destination:                                    # Heb even toegevoegd dat als het vliegtuig niet tussen airports gaat fixed cost ook 0 zijn, want staat stil
        C_fixed = 0
    C_time = C_t * (d[orgin, destination] / Speed)
    C_fuel = ((C_f * 1.42 / 1.5) * d[orgin, destination])

    calculate_operating_cost = (C_fixed + C_time + C_fuel) #(Toegevoegd om het weer runnend te krijgen)

    return calculate_operating_cost

print(f'Operation cost €{operating_costs(2, 3, 0):.2f}')
            
#TEST OM ALLE OPERATING COST TE PRINTEN (als de vliegtuigen groter worden wordt ook cost hoger ;), ook gecheckt met oude manier, zijn gelijk.
#k = 1 
#df_cost = pd.DataFrame(Ck_ij[:, :, k], index=airports, columns=airports)
#print(df_cost.round(0))

#revenue function
def revenue_function(orgin, destination, flow):
    revenue = (5.9 * d[orgin, destination]**(-0.76) + 0.043)* d[orgin, destination] * flow 
    return revenue


print(f'Revenue €{revenue_function(2, 3, 45*0.8):.2f}')

# DACHT MAAK ER EEN DICONARY VAN WANT DAN KAN JE UITEINDELIJK DAAR DE GEVLOGEN REIZIGERS VANAF TREKKEN HAD DAT EERST OOK NIET
demand_dic = {}                                                 # Create dictonary from demand. Then will be easier to update demand later
for i in range(n):
    for j in range(n):
        demand_dic[(i,j)] = {}
        for t in range(24):
            demand_dic[(i,j)][t] = D[i,j] * H[i,t]
            
            

def potential_flow(i, j, hour, cap, demand):
    flow = 0
    for h in [hour, hour-1, hour-2]:
        if h >= 0:
            flow += demand[(i,j)][h]
    return min(cap, flow)

def dynamic_programming(aircraft_type, demand_dic):

    total_steps = 24 * 10  # 6-minute steps
    n_airports = len(airports)

    profit_matrix = np.full((n_airports, total_steps + 1), -1e9)
    action_matrix = np.full((n_airports, total_steps), -1, dtype=int)

    # Terminal condition
    for j in range(n_airports):
        if j == hub_index:
            profit_matrix[j, total_steps] = 0
        else:
            profit_matrix[j, total_steps] = -1e9

    # Maak lokale kopie van demand voor deze DP-run
    demand_copy = copy.deepcopy(demand_dic)

    # Backward DP
    for t in range(total_steps - 1, -1, -1):
        current_hour = t // 10
        current_time = t * 6  # minuten sinds middernacht

        for i in range(n_airports):
            best_value = -1e9
            best_action = -1

            for j in action_possible(i, aircraft_type):

                # Optie 1: blijf
                if j == i:
                    value = 0

                # Optie 2: vlieg
                else:
                    bt = block_time(i, j, aircraft_type)
                    arrival_time = current_time + bt
                    arrival_step = math.ceil(arrival_time / 6)

                    if arrival_step > total_steps:
                        continue

                    future_profit = profit_matrix[j, arrival_step]

                    cap = LF * s[aircraft_type]
                    flow = potential_flow(i, j, current_hour, cap, demand_copy)
                    revenue = revenue_function(i, j, flow)
                    cost = operating_costs(i, j, aircraft_type)
                    profit = revenue - cost

                    if profit <= 0:
                        continue  # verliesmakende vlucht overslaan

                    value = profit + future_profit

            # Update best value/action
                if value > best_value:
                    best_value = value
                    best_action = j

            profit_matrix[i, t] = best_value
            action_matrix[i, t] = best_action

            # Update vraag direct als DP kiest voor vliegen (lokale kopie)
            if best_action != -1 and best_action != i:
                dep_hour = t // 10
                remaining_cap = potential_flow(i, best_action, dep_hour, LF * s[aircraft_type], demand_copy)
                for h in [dep_hour, dep_hour-1, dep_hour-2]:
                    if h >= 0 and remaining_cap > 0:
                        take = min(demand_copy[(i, best_action)][h], remaining_cap)
                        demand_copy[(i, best_action)][h] -= take
                        remaining_cap -= take

    return action_matrix, profit_matrix


def update_demand(demand, routing, aircraft_type):
    new_demand = copy.deepcopy(demand)
    cap = LF * s[aircraft_type]

    for (i, j, dep_t, arr_t) in routing:
        # Bereken flow op basis van originele vraag
        dep_hour = dep_t // 10
        remaining_cap = cap

        for h in [dep_hour, dep_hour - 1, dep_hour - 2]:
            if h >= 0 and remaining_cap > 0:
                take = min(new_demand[(i, j)][h], remaining_cap)
                new_demand[(i, j)][h] -= take
                remaining_cap -= take

    return new_demand

def route_profit(routing, aircraft_type, demand):
    """
    Bereken de totale profit van een routing.
    routing: lijst van tuples (i, j, dep_t, arr_t)
    aircraft_type: index van het vliegtuigtype
    demand: actuele vraag per uur (dictionary)
    """
    total_profit = 0
    for (i, j, dep_t, arr_t) in routing:
        dep_hour = dep_t // 10
        cap = LF * s[aircraft_type]
        flow = potential_flow(i, j, dep_hour, cap, demand)
        revenue = revenue_function(i, j, flow)
        cost = operating_costs(i, j, aircraft_type)
        total_profit += (revenue - cost)
    return total_profit




# ===============================
# MAIN PROGRAM – MULTI-AIRCRAFT
# ===============================

TOTAL_STEPS = 24 * 10
hub_index = hub_index  # je hub (AMS)
LF = 0.8  # load factor

# Voorbeeld: aantal vliegtuigen per type
available_aircrafts = list(Fleet)  # aantal vliegtuigen per type

solution_dict = {}  # hier slaan we de geplande vliegtuigen op
iteration = 0

while any(v > 0 for v in available_aircrafts):

    profits = []
    routings = []

    for ac_type, available in enumerate(available_aircrafts):
        if available <= 0:
            profits.append(-1e9)  # geen beschikbaar vliegtuig
            routings.append([])
            continue

        # ===== 1. DP uitvoeren =====
        action_matrix, profit_matrix = dynamic_programming(ac_type, demand_dic)

        # ===== 2. Route reconstrueren =====
        routing = []
        i = hub_index
        t = 0

        local_demand = copy.deepcopy(demand_dic)

        while t < TOTAL_STEPS:
            j = int(action_matrix[i, t])
            if j == -1:
                break

            if j != i:
                dep_t = t
                bt = block_time(i, j, ac_type)
                arr_t = t + math.ceil(bt / 6)

                dep_hour = dep_t // 10
                cap = LF * s[ac_type]
                flow = potential_flow(i, j, dep_hour, cap, local_demand)

                # Sla vlucht over als er geen passagiers zijn of verlies
                revenue = revenue_function(i, j, flow)
                cost = operating_costs(i, j, ac_type)
                if flow <= 0 or revenue - cost <= 0:
                    t += 1
                    continue

                routing.append((i, j, dep_t, arr_t))
                t += math.ceil(bt / 6)
                i = j
            else:
                t += 1

        # ===== 3. Bereken echte profit =====
        if len(routing) == 0:
            profit = -1e9  # geen winstgevende route
        else:
            profit = route_profit(routing, ac_type, demand_dic)

        profits.append(profit)
        routings.append(routing)

    # ===== 4. Stoppen als geen winstgevende routes meer =====
    if max(profits) <= 0:
        break

    # ===== 5. Kies het meest winstgevende vliegtuig =====
    best_idx = np.argmax(profits)
    solution_dict[f'Aircraft {iteration}'] = {
        'Aircraft type': best_idx,
        'Profit': profits[best_idx],
        'Routing': routings[best_idx],
        'Utilisation time': sum(block_time(i, j, best_idx) for (i,j,_,_) in routings[best_idx])
    }

    # ===== 6. Update demand =====
    demand_dic = update_demand(demand_dic, routings[best_idx], best_idx)

    # ===== 7. Verminder beschikbaar aantal voor dit type =====
    available_aircrafts[best_idx] -= 1
    iteration += 1


# ===============================
print("="*50)
print("Flight Schedule, Passengers & Profit per Aircraft")
print("="*50)

for idx, (ac_name, route_info) in enumerate(solution_dict.items()):
    ac_type = route_info['Aircraft type']          # type van vliegtuig
    routing = route_info['Routing']                # lijst van vluchten (i,j,dep_t,arr_t)
    profit = route_info['Profit']                  # totale profit
    total_bt = sum(block_time(i, j, ac_type) for (i, j, _, _) in routing)  # totale blocktime in minuten

    print(f"=== Aircraft {idx} ({ac_type}) ===")
    for (i, j, dep_t, arr_t) in routing:
        dep_hour, dep_min = divmod(dep_t * 6, 60)
        arr_hour, arr_min = divmod(arr_t * 6, 60)
        
        # Bereken aantal passagiers
        dep_hour_int = dep_t // 10
        cap = LF * s[ac_type]
        passengers = potential_flow(i, j, dep_hour_int, cap, demand_dic)
        
        print(f"{airports[i]} → {airports[j]} dep: {int(dep_hour):02d}:{int(dep_min):02d} arr: {int(arr_hour):02d}:{int(arr_min):02d} passengers: {int(passengers)}")
    
    print(f"Block time (h): {total_bt / 60:.2f}")
    print(f"Profit: {profit:.2f}")
    print("-"*50)




