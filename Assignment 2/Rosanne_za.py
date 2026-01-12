import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import * 
import openpyxl
import math
import copy

wb = load_workbook("Assignment 2/DemandGroup40.xlsx", data_only=True)
ws = wb.active

icao_row = 5    
lat_row = 6    
lon_row = 7    
start_col = 3
runway_row = 8
runways = []
hub_index = 2

RE = 6371.0                                     # Radius Earth
f = 1.42                                        # Fuel cost constant 

airports = []
latitudes = []
longitudes = []


col = start_col

while True:
    icao = ws.cell(row=icao_row, column=col).value
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    runway = ws.cell(row=runway_row, column=col).value
    if icao is None:
        break
    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    runways.append(float(runway))
    col += 1

#DISTANCE FORMULA

def distance(phi_i, lam_i, phi_j, lam_j):
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 +np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))




#DEMAND IMPORT
demand_start_row = icao_row + 7
demand_start_col = start_col 

n = len(airports)
D = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        cell = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(cell) if cell is not None else 0
        except:
            D[i, j] = 0


#FLEET IMPORT. --> Heel anders gedaan dan inladen van andere data: aanpassen zodat het een geheel wordt
wb2 = openpyxl.load_workbook("Assignment 2/FleetType.xlsx", data_only=True)
sheet2 = wb2.active

aircraft_names = [cell.value for cell in sheet2[1][1:]if cell.value is not None]
n_ac = len(aircraft_names)
data = {}

for row in sheet2.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue  
    parameter_name = row[0]
    values = row[1:1 + n_ac]  
    data[parameter_name] = values

df_aircraft = pd.DataFrame(data, index=aircraft_names) #vooral hier anders met aanmaken nieuwe dataset

print(df_aircraft)

#HOUR COEFFICIENT
wb = openpyxl.load_workbook("Assignment 2/HourCoefficients.xlsx", data_only=True)
ws = wb.active

start_row = 3        #AMSTERDAM
start_col = 4        #Hour 0 
n = len(airports)    #LENGTE RIJEN
T = 24               #AANTAL UREN (D t/m AA)

H = np.zeros((n, T))

for i in range(n):
    for t in range(T):
        cell = ws.cell(row=start_row + i, column=start_col + t).value
        try:
            H[i, t] = float(cell) if cell is not None else 0
        except:
            H[i, t] = 0

# DATA IMPORT
N = range(len(airports))                    # Set of airports; i, j in N
K = range(len(df_aircraft))                 # Set of aircrafts; k in K
ac = len(df_aircraft)                       # Total amount of aircrafts


n = len(airports) #distance
d = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        d[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])


y = np.zeros((n, n))
for i in N:                                 #YIELD
    for j in N:
        if d[i, j] > 0:
            y[i, j] = (5.9 * d[i, j]**(-0.76) + 0.043)
        else:
            y[i, j] = 0 

v = np.zeros(ac)                            #SPEED OF AIRPLANES (k)
for k in K:
    v[k] = df_aircraft['Speed [km/h]'].iloc[k]
    
s = np.zeros(ac)                            #SEATS
for k in K:
    s[k] = df_aircraft['Seats'].iloc[k]

ra = np.zeros(ac)                           #RANGE OF AIRCRAFT
for k in K:
    ra[k] = df_aircraft['Maximum Range [km]'].iloc[k]

RAC = np.zeros(ac)                          #RUNWAY REQUIREMENT DEPARTURE
for k in K:
    RAC[k] = df_aircraft['Runway Required [m]'].iloc[k]

RAP = np.zeros(n)                           #RUNWAY REQUIREMENT ARRIVAL
for j in N:
    RAP[j] = runways[j]

TAT = np.zeros(ac)                                                      #TURN AROUND TIMES IN MIN
for k in K:
    TAT[k] = df_aircraft['Average TAT [min]'].iloc[k]

Fleet = np.zeros(ac)                                                      #Fleettypes
for k in K:
    Fleet[k] = df_aircraft['Fleet'].iloc[k]
print(f'De Fleet is {Fleet}')

#COSTS 

cl = np.zeros(ac)                           #COSTS WEEKLY LEASE
for k in K:
    cl[k] = df_aircraft['Lease Cost [€/day]'].iloc[k]
    

C_fix = df_aircraft['Fixed Operating Cost (Per Fligth Leg)  [€]'].values
CT    = df_aircraft['Cost per Hour'].values
CF    = df_aircraft['Fuel Cost Parameter'].values
v     = df_aircraft['Speed [km/h]'].values



a = {}                                  #BIG M CONSTRAIN FOR RUNWAYS
for i in N:
    for j in N:
        for k in K:
            if d[i,j] <= ra[k]:
                a[i,j,k] = 10000
            else:
                a[i,j,k] = 0

LF = 0.80               # van 0.75 naar 0.80



#======== constraints airports functie (runway, range) =========


    #stage = index van luchthaven
    #required runway for aircraft < availible runway AND range arcraft 
def  action_possible(stage, aircraft_type):  
    possible_destinations = []

    #optie 1: aircraft is at hub 
    if stage == hub_index:
        candidate_destinations = range(len(airports)) # (choose other airport (can be hub))
    #optie 2: aircraft is not at hub 
    else:
        candidate_destinations = [stage, hub_index] # (stay or back to hub)


    for destination in candidate_destinations:
            if (RAC[aircraft_type] < RAP[destination] and ra[aircraft_type] >= d[stage, destination]):
                possible_destinations.append(destination)
    return(possible_destinations)

#TEST als je vanaf hub kijkt heb je met vliegtuig 1 20 opties
print(f'test action: {action_possible(2, 1)}')



#======== Time calculation --> function to calculate blocking time of one flight ========== 

def block_time(origin, destination, aircraft_type):
    if origin == destination: #als vliegtuig blijft staan, geen blocking time
        BT = 0
    else:
        BT = 15 + (d[origin, destination] / v[aircraft_type]) * 60 + TAT[aircraft_type] + 15
    return BT

#TEST berekenen blocking time: van ams naar ams met aircraft 1 
print(f'test bt: {block_time(2, 4, 1):.2f} min, {block_time(2, 4, 1) / 60:.2f} uur')




#===========Scheduling horizon: time steps ========

#time step has to be converted to a time 00:06, 00:12 etc
def timestep_converting(timestep):
    step = 6 # 6 min per step
    steps_day = 24 * 10
    total_minutes = (timestep % steps_day) * step # tijdstap binnen de dag
    
    hours = total_minutes // 60
    minutes = total_minutes % 60
    time = f"{hours:02d}:{minutes:02d}"
    return hours, minutes 

#TEST convert timestep to tijdstip voor timetable: LET OP: step 1 = 00:00 
print(f'test time= {timestep_converting(3)}')

#======== Operating cost function ==========

def operating_costs(origin, destination, aircraft_type):
    
    C_t = CT[aircraft_type]
    C_f = CF[aircraft_type]
    Speed = v[aircraft_type]

    C_fixed = C_fix[aircraft_type]
    if origin == destination:                                    # Heb even toegevoegd dat als het vliegtuig niet tussen airports gaat fixed cost ook 0 zijn, want staat stil
        C_fixed = 0
    C_time = C_t * (d[origin, destination] / Speed)
    C_fuel = ((C_f * 1.42 / 1.5) * d[origin, destination])

    calculate_operating_cost = (C_fixed + C_time + C_fuel)      # (Toegevoegd om het weer runnend te krijgen)

    return calculate_operating_cost

print(f'Operation cost €{operating_costs(2, 2, 0):.2f}')


            
#TEST OM ALLE OPERATING COST TE PRINTEN (als de vliegtuigen groter worden wordt ook cost hoger ;), ook gecheckt met oude manier, zijn gelijk.
#k = 1 
#df_cost = pd.DataFrame(Ck_ij[:, :, k], index=airports, columns=airports)
#print(df_cost.round(0))

#========== revenue function =========

def revenue_function(origin, destination, flow):
    if origin == destination:
        revenue = 0
    else:   
        revenue = (5.9 * d[origin, destination]**(-0.76) + 0.043)* d[origin, destination] * flow 
    return revenue


print(f'Revenue €{revenue_function(2, 3, 45*0.8):.2f}')


#=========== Demand function ========== ROSANNE 
dem_hour = np.zeros((n,n, 24))

for i in range(n):
    for j in range(n):
        for t in range(24):
            dem_hour[i,j,t] = D[i,j] * H[i,t]


#=========== Capture demand function ==========

def capture_demand(origin, destination, timestep, aircraft_type, dem_hour, verbose=False):
    hour, minute = timestep_converting(timestep)
    capacity = 0.8 * s[aircraft_type]
    remaining_capacity = capacity
    flow = 0

    #update hours 
    for h in [hour, hour-1, hour-2]:
        if h < 0:
            continue
        
        demand_left = dem_hour[origin, destination, h]
        if demand_left <= 0:
            continue

        demand_taken = min(demand_left, remaining_capacity)

        dem_hour[origin, destination, h] -= demand_taken
        remaining_capacity -= demand_taken
        flow += demand_taken

        # if verbose:
        #     print(f"Hour {h}: taken={demand_taken:.1f}, remaining={dem_hour[origin, destination, h]:.1f}")

        if remaining_capacity <= 0:
            break

    # if verbose:
    #     print(f"Total passengers captured for {airports[origin]} -> {airports[destination]} at timestep {timestep}: {flow:.1f}\n")

    return flow, dem_hour


def dynamic_programming(aircraft_type, demand_dic):
    total_steps = 24 * 10           # Amount of steps in the day

    profit_matrix = np.zeros((len(airports), total_steps))      # Storage space for profit
    action_matrix = np.zeros((len(airports), total_steps))      # Storage of action taken

    for i in range(total_steps):
        current_time_step = total_steps - (i+1)                 # Starts at the end, but the very last is already determined
        current_time = current_time_step * 6

        # if current_time_step % 10 == 0:
        #     print(current_time_step)                            # Control for if the function is going trough all time steps
        
        for j in range(len(airports)):
            best_action = -1                                    # Set base for action value, if action is  chosen it will get a positive value

            if j == hub_index and i == 0:
                best_solution = 0                               # For start position best solution is 0, is given
            else:
                best_solution = -10e5                           # Base every where big negative so it can only improve

            for k in action_possible(j, aircraft_type):

                if k == j:
                    if current_time_step < total_steps - 1:
                        current_solution = profit_matrix[j][current_time_step + 1]      # Previous profit is same as current
                    else:
                        current_solution = -10e5

                else:
                     fligth_duration = block_time(j, k, aircraft_type)
                     arrival_time = current_time + fligth_duration
                     arrival_time_step = math.ceil(arrival_time / 6) - 1
                     if arrival_time_step >= total_steps:
                        future_profit = 0
                     else:
                        future_profit = profit_matrix[k][arrival_time_step]
                     
                     operation_cost = operating_costs(j, k, aircraft_type)
                     current_hour, current_min = timestep_converting(current_time_step)
                     capacity = 0.8 * s[aircraft_type]                  #load factor

                     expected_demand = 0
                     for h in [current_hour, current_hour - 1, current_hour - 2]:
                        if h >= 0:
                            expected_demand += demand_dic[j, k, h]

                     flow = min(capacity, expected_demand)

                     revenue = revenue_function(j, k, flow)
                     profit = revenue - operation_cost

                     current_solution = profit + future_profit

                if current_solution > best_solution:
                    best_solution = current_solution
                    best_action = k

            action_matrix[j, current_time_step] = best_action
            profit_matrix[j, current_time_step] = best_solution


    return action_matrix, profit_matrix

print(f'Poging {dynamic_programming(1, dem_hour)}')



#=========== Schedule function ==========

def schedule(aircraft_type, action_matrix, profit_matrix, dem_hour):
    optimal_route   = [(0, hub_index)]
    time_step       = 0
    current_airport = hub_index
    op_time         = 0
    total_steps     = 24 * 10

    flows_per_segment = []

    # werk met een diepe kopie van demand
    remaining_demand = copy.deepcopy(dem_hour)

    while time_step < total_steps:
        next_airport = int(action_matrix[current_airport][time_step])

        # stopconditie: geen geldige actie
        if next_airport < 0 or next_airport >= len(airports):
            break

        # ===== STAY =====
        if next_airport == current_airport:
            next_time_step = time_step + 1

        # ===== FLY =====
        else:
            bt = block_time(current_airport, next_airport, aircraft_type)
            step_increase = math.ceil(bt / 6)
            next_time_step = time_step + step_increase

            if next_time_step >= total_steps:
                break

            op_time += bt

            # echte demand-afboeking
            flown, remaining_demand = capture_demand(
                current_airport,
                next_airport,
                time_step,
                aircraft_type,
                remaining_demand,
                verbose=True
            )

            flows_per_segment.append(flown)
            optimal_route.append((next_time_step, next_airport))
            # print(
            #     f"FLIGHT | AC {aircraft_type} | "
            #     f"{airports[current_airport]} -> {airports[next_airport]} | "
            #     f"t={time_step} | bt={bt:.1f} min | flow={flown:.0f}")

        # update state
        current_airport = next_airport
        time_step = next_time_step

    # totale profit = DP-profit minus leasekosten
    profit = profit_matrix[hub_index][0] - cl[aircraft_type]

    min_block_time = 6 * 60  # 360 minuten

    # check minimum block time
    if op_time < min_block_time:
        profit = -1e9  # maak oplossing extreem ongunstig

    return optimal_route, profit, op_time, flows_per_segment, remaining_demand


# Pre process
demand = dem_hour
available_ac = list(Fleet)
solution_dict = {}
iteration = 0

while any(value > 0 for value in Fleet):
    optimal_route = {aircraft: [] for aircraft in K}
    optimal_route = {aircraft: [] for aircraft in K}
    profits = np.zeros(3)
    ut_time = np.zeros(3)

    for k in [i for i, value in enumerate(available_ac) if value > 0]:
        state, action = dynamic_programming(k, demand)
        if k == 0:
            print(state[1][0])
        optimal_route[k], profits[k], ut_time[k], flow, x = schedule(k, state, action, demand)
    
    print("Iteration: {}" .format(iteration), profits, available_ac)

    for k in K:
        if ut_time[k] < 6 * 60:
            profits[k] = 0

    print("Iteration: {}" .format(iteration), profits, available_ac)

    if not any(value > 0 for value in profits):
        break

    k = np.argmax(profits)

    solution_dict["Route {}" .format(iteration)] = {"Aircraft type": K[k],
                                                  "Profit": profits[k],
                                                  "Utilisation time": ut_time[k],
                                                  "Routing": optimal_route[k]}
    
    print("Revenue:", action[hub_index][0],
      "Lease:", cl[k],
      "Net:", profits[k])

    
    optimal_route[k], profits[k], ut_time[k], flow, new_demand = schedule(k, state, action, demand)

    demand = new_demand

    available_ac[k] -= 1
    iteration += 1

# #=========== Build timetable with revenue, cost, profit ==========

# def timestep_to_time(timestep):
#     minutes = timestep * 6
#     h = minutes // 60
#     m = minutes % 60
#     return f"{int(h):02d}:{int(m):02d}"


# def build_timetable(all_schedules):
#     """
#     all_schedules: list of dicts with keys:
#         - aircraft_type
#         - route
#         - flows
#     """
#     timetable = []
#     flight_id = 1
#     aircraft_used_count = {k: 0 for k in K}

#     for sched in all_schedules:
#         ac_type = sched["aircraft_type"]
#         route = sched["route"]
#         flows = sched["flows"]

#         if len(route) > 1:
#             aircraft_used_count[ac_type] += 1

#         flow_idx = 0
#         for i in range(len(route) - 1):
#             dep_step, origin = route[i]
#             arr_step, destination = route[i + 1]

#             if origin == destination:
#                 continue  # skip idle steps

#             flow = flows[flow_idx]
#             flow_idx += 1

#             # Bereken kosten, revenue en winst
#             rev = revenue_function(origin, destination, flow)
#             cost = operating_costs(origin, destination, ac_type)
#             profit = rev - cost

#             timetable.append({
#                 "Flight": flight_id,
#                 "Aircraft type": ac_type,
#                 "From": airports[origin],
#                 "To": airports[destination],
#                 "Departure": timestep_to_time(dep_step),
#                 "Arrival": timestep_to_time(arr_step),
#                 "Passengers": int(round(flow)),
#                 "Distance": d[origin, destination],
#                 "Revenue (€)": rev,
#                 "Cost (€)": cost,
#                 "Profit (€)": profit
#             })

#             flight_id += 1

#     return pd.DataFrame(timetable), aircraft_used_count


# def print_timetable(df):
#     print("\n===== DAILY FLIGHT TIMETABLE =====\n")
#     print(df.to_string(index=False))


# #=========== Plan all aircraft and build timetable ==========
# all_schedules = []

# for k in K:
#     action_matrix, profit_matrix = dynamic_programming(k, dem_hour)
#     route, profit, op_time, flows = schedule(k, action_matrix, profit_matrix, dem_hour)

#     # skip onrendabele aircraft
#     if profit < 0:
#         continue

#     # update demand voor volgende aircraft
#     for i, (t, airport) in enumerate(route[:-1]):
#         flown, dem_hour = capture_demand(route[i][1], route[i+1][1], t, k, dem_hour)

#     all_schedules.append({
#         "aircraft_type": k,
#         "route": route,
#         "flows": flows,
#         "profit": profit,
#         "block_time": op_time
#     })

# # Build timetable + aircraft usage
# timetable_df, aircraft_used_count = build_timetable(all_schedules)

# # Print timetable
# print_timetable(timetable_df)

# # Print totale revenue, kosten en winst
# total_revenue = timetable_df["Revenue (€)"].sum()
# total_cost = timetable_df["Cost (€)"].sum()
# total_profit = timetable_df["Profit (€)"].sum()

# print("\n===== TOTALS =====")
# print(f"Totale revenue: {total_revenue:.0f} €")
# print(f"Totale kosten: {total_cost:.0f} €")
# print(f"Totale winst: {total_profit:.0f} €\n")

# print("Aircraft gebruikt per type:")
# for k, count in aircraft_used_count.items():
#     print(f"Type {k+1}: {count} aircraft(s)")
