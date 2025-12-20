from gurobipy import * 
from openpyxl import * 
import openpyxl
from time import *
import numpy as np
import math
import matplotlib.pyplot as plt
import pandas as pd
import statsmodels.api as sm

#IMPORT ALL THE EXCEL FILES FOR ASSIGNMENT 1A

wb = load_workbook("pop.xlsx", data_only=True)
ws = wb["General"]

wb_codes = load_workbook("Airport_names.xlsx", data_only=True)
ws_codes = wb_codes.active

city_to_icao = {}
country_to_icao = {}

row = 1
while True:                                                             #Give everything a good name and couple them
    city = ws_codes.cell(row=row, column=1).value
    country = ws_codes.cell(row=row, column=2).value
    icao = ws_codes.cell(row=row, column=3).value

    if city is None:
        break

    city_to_icao[city] = icao
    country_to_icao[country] = icao

    row += 1

population_2021_dict = {}
population_2024_dict = {}
gdp_2021_dict = {}
gdp_2024_dict = {}

row = 4

#GET THE RIGHT INFORMATION FROM THE EXCELFILES

while True:
    city = ws.cell(row=row, column=1).value        # kolom A
    pop2021 = ws.cell(row=row, column=2).value     # kolom B
    pop2024 = ws.cell(row=row, column=3).value     # kolom C
    gdp2021 = ws.cell(row=row, column=6).value     # kolom F
    gdp2024 = ws.cell(row=row, column=7).value     # kolom G

    if city is None:
        break

    if city in city_to_icao:
        icao = city_to_icao[city]
        population_2021_dict[icao] = pop2021
        population_2024_dict[icao] = pop2024
        gdp_2021_dict[icao] = gdp2021
        gdp_2024_dict[icao] = gdp2024

    row += 1

wb = load_workbook("DemandGroup40.xlsx", data_only=True)
ws = wb.active

#INFORMATION FOR ASSIGNMENT 1B
icao_row = 5    
lat_row = 6    
lon_row = 7    
start_col = 3
runway_row = 8
slots_row  = 9
runways = []
slots = []
hub_index = 2                                   # Amsterdam

RE = 6371.0                                     # Radius Earth
f = 1.42                                        # Fuel cost constant 

airports = []
latitudes = []
longitudes = []

col = start_col
while True:
    icao = ws.cell(row=icao_row, column=col).value
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    runway = ws.cell(row=runway_row, column=col).value
    slot = ws.cell(row=slots_row, column=col).value
    if icao is None:
        break
    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    runways.append(float(runway))
    if slot == '-':
        slots.append(float('inf'))
    else:
        slots.append(float(slot))
    col += 1

#DISTANCE FORMULA

def distance(phi_i, lam_i, phi_j, lam_j):
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 +np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))


n = len(airports)
dij = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        dij[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])


demand_start_row = icao_row + 8 
demand_start_col = start_col 


#BUILDING UP D (IJ)

n = len(airports)
D = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        cell = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(cell) if cell is not None else 0
        except:
            D[i, j] = 0

n = len(airports)
rows = []

for i in range(n):        
    for j in range(n):
        if i == j:
            continue                            # If orgin and destionation is the same, go on

        Dij = D[i, j]
        if Dij <= 0:
            continue                            # It there is no demand, go on

        i_icao = airports[i]                    #Collectiong ICAO code
        j_icao = airports[j]

        row = {
            "D": Dij,
            "pop": population_2021_dict[i_icao] * population_2021_dict[j_icao],
            "gdp": gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao],
            "fd": f * dij[i, j]
        }
        rows.append(row)

df = pd.DataFrame(rows)

#LINEAR REGRESSION

df['lnD']    = np.log(df['D'])                  
df['ln_pop'] = np.log(df['pop'])
df['ln_gdp'] = np.log(df['gdp'])
df['ln_fd']  = np.log(df['fd'])


X = df[['ln_pop', 'ln_gdp', 'ln_fd']]           # Independent variables
X = sm.add_constant(X)                          # This creates ln(k)
y = df['lnD']                                   # Dependent variable

model = sm.OLS(y, X).fit()                      # Ordinary least squares
#print(model.summary())


a = model.params['const']            
b1 = model.params['ln_pop']         
b2 = model.params['ln_gdp']

beta3 = model.params['ln_fd']        
b3 = -beta3                          

k1 = np.exp(a)                       

#CREATE PREDICITION 2021

D_pred = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred[i, j] = 0  
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product = population_2021_dict[i_icao] * population_2021_dict[j_icao]
        gdp_product = gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred[i, j] = k1 * (pop_product**b1) * (gdp_product**b2) * ((fd_ij)**(beta3))


# DIFFERENCE MATRIX
D_diff = D_pred - D  


# POP EN GDP 2021 AND 2024 FOR GROWFACTORS
population_2026_dict = {}
pop_g_list = []                                         #Average pop g list

for icao in population_2021_dict:
    pop_2021 = population_2021_dict[icao]
    pop_2024 = population_2024_dict[icao]

    g = (pop_2024 / pop_2021)**(1/3)                   #Annual growth
    pop_g_list.append(g)
    pop_2026 = pop_2024 * (g)**2
    population_2026_dict[icao] = pop_2026              #Prediction 2026

gdp_2026_dict = {}
gdp_g_list = []

for icao in gdp_2021_dict:
    gdp_2021 = gdp_2021_dict[icao]
    gdp_2024 = gdp_2024_dict[icao]

    g = (gdp_2024 / gdp_2021)**(1/3)
    gdp_g_list.append(g)
    
    gdp_2026 = gdp_2024 * (g)**2
    gdp_2026_dict[icao] = gdp_2026


avg_gdp_g = sum(gdp_g_list) / len(gdp_g_list)
#print(f"Gemiddelde jaarlijkse groeifactor GDP: {avg_gdp_g:.4f}")

avg_pop_g = sum(pop_g_list) / len(pop_g_list)
#print(f"Gemiddelde jaarlijkse groeifactor populatie: {avg_pop_g:.4f}")

#PREDICTION 2026
D_pred_26 = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred_26[i, j] = 0  
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product_26 = population_2026_dict[i_icao] * population_2026_dict[j_icao]
        gdp_product_26 = gdp_2026_dict[i_icao] * gdp_2026_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred_26[i, j] = k1 * (pop_product_26**b1) * (gdp_product_26**b2) * ((fd_ij)**(beta3))


D = np.array(D_pred_26)

df['lnD_pred'] = model.predict(X)

df['D_pred'] = np.exp(df['lnD_pred'])
df['D_actual'] = np.exp(df['lnD'])

x = df['D_actual']
y = df['D_pred']



### IMPLEMENTATION ASSIGNMENT 1B 

# LOAD AIRCRAFT DATA 
wb2 = openpyxl.load_workbook("AircraftData.xlsx")
sheet2 = wb2.active  
aircraft_names = [cell.value for cell in sheet2[1][1:]]  
data = {}
for row in sheet2.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        continue
    
    parameter_name = row[0]
    values = row[1:]
    
    data[parameter_name] = values

# TRANSPOSE AIRCFRAFT BECOMES ROWS
df_aircraft = pd.DataFrame(data, index=aircraft_names)
df_aircraft = df_aircraft.drop(columns=["Aircraft Characteristics"])


# DATA IMPORT
N = range(len(airports))                    # Set of airports; i, j in N
K = range(len(df_aircraft))                 # Set of aircrafts; k in K
ac = len(df_aircraft)                       # Total amount of aircrafts

# CREATE PARAMETERS
q = np.zeros((n, n))                        # Demand between orgin i and destination j
for i in N:
    for j in N:
        q[i, j] = D_pred_26[i, j]

d = np.zeros((n, n))                        # Distance between i and j
for i in N:
    for j in N:
        d[i, j] = dij[i, j]


y = np.zeros((n, n))
for i in N:                                 #YIELD
    for j in N:
        if d[i, j] > 0:
            y[i, j] = (5.9 * d[i, j]**(-0.76) + 0.043)
        else:
            y[i, j] = 0   


s = np.zeros(ac)                            #SEATS
for k in K:
    s[k] = df_aircraft['Seats'][k]

v = np.zeros(ac)                            #SPEED OF AIRPLANES (k)
for k in K:
    v[k] = df_aircraft['Speed [km/h]'][k]


ra = np.zeros(ac)                           #RANGE OF AIRCRAFT
for k in K:
    ra[k] = df_aircraft['Maximum range [km]'][k]

RAC = np.zeros(ac)                          #RUNWAY REQUIREMENT DEPARTURE
for k in K:
    RAC[k] = df_aircraft['Runway required [m]'][k]

RAP = np.zeros(n)                           #RUNWAY REQUIREMENT ARRIVAL
for j in N:
    RAP[j] = runways[j]

cl = np.zeros(ac)                           #COSTS WEEKLY LEASE
for k in K:
    cl[k] = df_aircraft['Weekly lease cost [€]'][k]
   

C = np.zeros(ac)                            #COST FIXED OPERATINGH
for k in K:
    C[k] = df_aircraft['Fixed operating cost C_X [€]'][k]

CT = np.zeros(ac)                           #TIME COST PARAMETER
for k in K:
    CT[k] = df_aircraft['Time cost parameter C_T [€/hr]'][k]

C_Tij = np.zeros((n, n, ac))                #TIME COST PER LEG
for k in K:
    C_Tij[ :, :,k] = CT[k] * (d/ v[k])
    

CF = np.zeros(ac)                           #FUEL COST PARAMETER
for k in K:
    CF[k] = df_aircraft['Fuel cost parameter C_F'][k]

C_Fij = np.zeros((n, n, ac))                #FUEL COST PER LEG
for k in K:
    C_Fij[:, :, k] = ((CF[k] * 1.42) / 1.5) * d


Ck_ij = np.zeros((n, n, ac))                #tOTAL OPERARION COST PER LEG (WITH DISCOUNT OF 30%)
for k in K:
    for i in N:
        for j in N:
            Ck_ij[i, j, k] = (C[k] + C_Tij[i, j, k] + C_Fij[i, j, k]) * 0.7
            

a = {}                                  #BIG M CONSTRAIN FOR RUNWAYS
for i in N:
    for j in N:
        for k in K:
            if d[i,j] <= ra[k]:
                a[i,j,k] = 10000
            else:
                a[i,j,k] = 0

g = np.ones(n)                      #HUB IS AMSTERDAM 
g[hub_index] = 0 

LF = 0.75                                                             # LOADFACTOR (GIVEN)

BT = 70                                                                # AVAILIBLE HOURS OF A PLANE IN A WEEK 

TS = np.zeros(n)                                                       # SLOTS 
for j in N:
    TS[j] = slots[j]

TAT = np.zeros(ac)                                                      #TURN AROUND TIMES IN HOURS
for k in K:
    TAT[k] = df_aircraft['Average TAT [mins]'][k]/60


from gurobipy import Model, GRB, quicksum

def main():
    model = Model("Model_1B")
    model.write("network_fleet_development.lp")
    model.setParam('TimeLimit', 10)

    x = model.addVars(N, N, name="x", vtype=GRB.INTEGER, lb=0)
    w = model.addVars(N, N, name="w", vtype=GRB.INTEGER, lb=0)
    z = model.addVars(N, N, K, name="z", vtype=GRB.INTEGER, lb=0)
    AC = model.addVars(K, name="AC", vtype=GRB.INTEGER, lb=0)

    #OBJECTIVE FUNCTION:
    
    Objective_1B = quicksum((y[i, j] * d[i, j] * (x[i, j] + w[i, j])) for i in N for j in N) - quicksum(Ck_ij[i, j, k] * z[i, j, k] for k in K for i in N for j in N) - quicksum(cl[k] * AC[k] for k in K)
    model.setObjective(Objective_1B, GRB.MAXIMIZE)

    #Passengers must be smaller than the demand
    for i in N:
        for j in N:
            model.addConstr(x[i, j] + w[i, j] <= q[i, j], name=f"demand_limit_{i}_{j}")

    #Passengers Transfer passenger if the hub is not their origin or destination
    for i in N:
        for j in N:
            model.addConstr(w[i,j]<= q[i,j] * g[i] * g[j],name=f"transfer")

    #Constrain that every flight have to go via the hub
    for i in N:
        for j in N:
            for k in K:
                model.addConstr(z[i, j, k] * g[i] * g[j] == 0, name=f"via_hub_{i}_{j}_{k}")


    #Balance incomming and outgoing flights
    for i in N:
        for k in K:
            lhs = quicksum(z[i, j, k]  for j in N)
            rhs = quicksum(z[j, i, k]  for j in N)
            model.addConstr(lhs == rhs, name=f"return_constrain_{i}_{k}")
    
    # Passengers smaller than amount of seats 
    for i in N:
        for j in N:
            lhs1 = x[i,j] + quicksum((w[i,m]*(1 - g[j])) for m in N) + quicksum((w[m,j]*(1 - g[i])) for m in N)
            rhs1 = quicksum((z[i, j, k]  * s[k] * LF) for k in K)
            model.addConstr(lhs1 <= rhs1, name=f"cap_constraint_{i}_{j}")


    #Duration constrain
    for k in K:
        lhs3 = quicksum((((d[i, hub_index] + d[hub_index, j])/ v[k]) + TAT[k] + (TAT[k] * 0.5 * (1 - g[j] ))) * z[i, j, k]  for i in N for j in N) 
        rhs3 = (BT * AC[k]) 
        model.addConstr(lhs3 <= rhs3, name=f"duration_constrain_{k}")

    # Range constrain
    for k in K:
        for i in N:
            for j in N:
                model.addConstr(z[i, j, k]  <= a[i, j, k], name=f"reach_{i}_{j}_{k}")

    #Runway constraints
    for k in K:
        for i in N:
            for j in N:
                model.addConstr(RAC[k] * z[i, j, k] <= RAP[i]* z[i, j, k] , name=f'Runway_dep_{i}_{j}_{k}') 

    for k in K:
        for i in N:
            for j in N:
                model.addConstr(RAC[k] * z[i, j, k]  <= RAP[j] * z[i, j, k] , name=f'Runway_arr_{i}_{j}_{k}') 

    #Timeslot constraints
    for j in N:
        model.addConstr(quicksum(z[i, j, k]  for i in N for k in K) <= TS[j], name=f'Time_slots_{j}')
 


    model.optimize()
    #model.write("Model_1B.sol")
    #model.write("Model_1B.lp")
    if model.status == GRB.OPTIMAL or model.status == GRB.TIME_LIMIT:
        total_profit = model.ObjVal  
        print("\nRESULTS:")
        print(f"Total profit: €{total_profit:.2f}")

        print("\nNumber of planes per type:")
        for k in K:
            print(f"{df_aircraft.index[k]}: {AC[k].X:.0f}")
            
        total_passengers = sum(x[i, j].X for i in N for j in N)
        print(f"\nTotal passengers depart/arrive from/at the hub: {total_passengers:.0f}")

        transfer_passagers = sum(w[i, j].X for i in N for j in N)
        print(f'\nNumber of transfer passengers: {transfer_passagers}')    

        for i in N:
            for j in N:
                for k in K:
                    if z[i, j, k].X > 0.5:  
                        total_passengers = z[i, j, k].X * s[k] * LF
                        print(f"{airports[i]} → {airports[j]} with {df_aircraft.index[k]} "
                            f"Number of flights: {z[i, j, k].X:.0f}, "
                            f"total passengers: {total_passengers:.0f}")

                        
        print("\nAverage flighttime per type:")
        for k in K:
            total_hours_k = sum(
                (((d[i, hub_index] + d[hub_index, j])/ v[k]) + TAT[k] + (TAT[k] * 0.5 * (1 -g[j]))) * z[i, j, k].X
                for i in N for j in N)

            num_aircraft = AC[k].X

            if num_aircraft > 0:
                hours_per_aircraft = total_hours_k / num_aircraft
                print(f"Type {k}: {hours_per_aircraft:.2f} uur per Airplane")
            else:
                print(f"Type {k}: Airplane not used")


        # print("\nLeft slots:")
        # unmet = np.zeros(n)
        # for j in N:
        #     sum_z = sum(z[i, j, k].X for i in N for k in K) 
        #     unmet[j] = TS[j] - sum_z
        #     print(f"{airports[j]}: {unmet[j]:.0f}")



        # print("\nMatrix x[i,j]:")
        # for i in N:
        #     row = ""
        #     for j in N:
        #         row += f"{x[i,j].X:8.1f} "
        #     print(row)

        # print("\nMatrix w[i,j]:")
        # for i in N:
        #     row = ""
        #     for j in N:
        #         row += f"{w[i,j].X:8.1f} "
        #     print(row)              
        
      
    else:
        print("No optimal solution found")

main()
