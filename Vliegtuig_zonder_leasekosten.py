from gurobipy import * 
from openpyxl import * 
import openpyxl
from time import *
import numpy as np
import math
import matplotlib.pyplot as plt
import pandas as pd
import statsmodels.api as sm

import sys
print(sys.executable)


wb = load_workbook("pop.xlsx", data_only=True)
ws = wb["General"]

population_2021 = []
gdp_2021 = []

population_2024 = []
gdp_2024 = []

row = 4  
while True:
    city = ws.cell(row=row, column=1).value          # kolom A
    pop2021 = ws.cell(row=row, column=2).value       # kolom B

    if city is None:
        break

    population_2021.append((city, pop2021))
    row += 1

row=4

while True:
    city = ws.cell(row=row, column=1).value          # kolom A
    pop2024 = ws.cell(row=row, column=3).value       # kolom C

    if city is None:
        break

    population_2024.append((city, pop2024))
    row += 1

row = 4  

while True:
    country = ws.cell(row=row, column=5).value       # kolom E
    gdp2021 = ws.cell(row=row, column=6).value       # kolom F

    if country is None:
        break  
    gdp_2021.append((country, gdp2021))
    row += 1

row = 4

while True:
    country = ws.cell(row=row, column=5).value       # kolom E
    gdp2024 = ws.cell(row=row, column=7).value       # kolom G

    if country is None:
        break  
    gdp_2024.append((country, gdp2024))
    row += 1


wb_codes = load_workbook("Airport_names.xlsx", data_only=True)
ws_codes = wb_codes.active

city_to_icao = {}
country_to_icao = {}

row = 1
while True:
    city = ws_codes.cell(row=row, column=1).value
    country = ws_codes.cell(row=row, column=2).value
    icao = ws_codes.cell(row=row, column=3).value

    if city is None:
        break

    city_to_icao[city] = icao
    country_to_icao[country] = icao

    row += 1

population_2021_dict = {}
population_2024_dict = {}
gdp_2021_dict = {}
gdp_2024_dict = {}

row = 4
while True:
    city = ws.cell(row=row, column=1).value        # kolom A
    pop2021 = ws.cell(row=row, column=2).value     # kolom B
    pop2024 = ws.cell(row=row, column=3).value     # kolom C
    gdp2021 = ws.cell(row=row, column=6).value     # kolom F
    gdp2024 = ws.cell(row=row, column=7).value     # kolom G

    if city is None:
        break

    if city in city_to_icao:
        icao = city_to_icao[city]
        population_2021_dict[icao] = pop2021
        population_2024_dict[icao] = pop2024
        gdp_2021_dict[icao] = gdp2021
        gdp_2024_dict[icao] = gdp2024

    row += 1

# print("Population:")
# print(population_2021_dict)

# print("GDP 2021:")
# print(gdp_2021_dict)

# print("Population 2024:")
# print(population_2024_dict)

# print("GDP 2024:")
# print(gdp_2024_dict)

wb = load_workbook("DemandGroup40.xlsx", data_only=True)
ws = wb.active

icao_row = 5    
lat_row = 6    
lon_row = 7    
start_col = 3
runway_row = 8
slots_row  = 9
runways = []
slots = []
hub_index = 2  # Amsterdam


airports = []
latitudes = []
longitudes = []

col = start_col
while True:
    icao = ws.cell(row=icao_row, column=col).value
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    runway = ws.cell(row=runway_row, column=col).value
    slot = ws.cell(row=slots_row, column=col).value
    if icao is None:
        break
    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    runways.append(float(runway))
    if slot == '-':
        slots.append(float('inf'))
    else:
        slots.append(float(slot))
    col += 1



RE = 6371.0 

def distance(phi_i, lam_i, phi_j, lam_j):
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 +np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))


n = len(airports)
dij = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        dij[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])


# print("Afstandsmatrix (km):")
# print("\t" + "\t".join(airports))
# for i, row in enumerate(dij):
#     print(airports[i], "\t" + "\t".join(f"{val:.2f}" for val in row))


demand_start_row = icao_row + 8 
demand_start_col = start_col 

n = len(airports)
D = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        cell = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(cell) if cell is not None else 0
        except:
            D[i, j] = 0

# print("Demandmatrix 2021 (week):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D[i,j]:.0f}" for j in range(n)))

# Determning B1, B2, B3 and k
f = 1.42   # fuel cost constant from assignment

n = len(airports)
rows = []

for i in range(n):        
    for j in range(n):
        if i == j:
            continue                    # If orgin and destionation is the same, go on

        Dij = D[i, j]
        if Dij <= 0:
            continue                    # It there is no demand, go on

        i_icao = airports[i]            #Collectiong ICAO code
        j_icao = airports[j]

        row = {
            "D": Dij,
            "pop": population_2021_dict[i_icao] * population_2021_dict[j_icao],
            "gdp": gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao],
            "fd": f * dij[i, j]
        }
        rows.append(row)

df = pd.DataFrame(rows)
# print(df)

# --- LOGS ---
df['lnD']    = np.log(df['D'])                  #logarithm to lineariz the model, dat doe je door ln van alles te nemen heeft google mij verteld
df['ln_pop'] = np.log(df['pop'])
df['ln_gdp'] = np.log(df['gdp'])
df['ln_fd']  = np.log(df['fd'])

# --- OLS REGRESSIE ---
X = df[['ln_pop', 'ln_gdp', 'ln_fd']]           # Independent variables
X = sm.add_constant(X)                          # This creates ln(k)
y = df['lnD']                                   # Dependent variable

model = sm.OLS(y, X).fit()                      # Ordinary least squares
# print(model.summary())

# --- PARAMETERS ---
a = model.params['const']            # ln(k)
b1 = model.params['ln_pop']          # Creating b1
b2 = model.params['ln_gdp']

beta3 = model.params['ln_fd']        # = -b3 (afstand in noemer)
b3 = -beta3                          # Because in function it is in the denominator

k1 = np.exp(a)                       # Making k again after it is a ln()

# print("\n--- GRAVITY MODEL PARAMETERS ---")
# print(f"k  = {k}")
# print(f"b1 = {b1:.4f}")
# print(f"b2 = {b2:.4f}")
# print(f"b3 = {b3:.4f}")

# Maak een lege matrix voor voorspelde demand
D_pred = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred[i, j] = 0  # geen vraag van een stad naar zichzelf
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product = population_2021_dict[i_icao] * population_2021_dict[j_icao]
        gdp_product = gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred[i, j] = k1 * (pop_product**b1) * (gdp_product**b2) * ((fd_ij)**(beta3))

# print("Voorspelde demand (gravity model):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D_pred[i,j]:.0f}" for j in range(n)))


# Verschilmatrix
D_diff = D_pred - D  # absoluut verschil

# print("Absoluut verschil (voorspeld - werkelijke vraag):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D_diff[i,j]:.0f}" for j in range(n)))

# POP EN GDP VOORSPELLEN 2026
population_2026_dict = {}
for icao in population_2021_dict:
    pop_2021 = population_2021_dict[icao]
    pop_2024 = population_2024_dict[icao]

    # jaarlijkse groeivoet
    g = (pop_2024 / pop_2021)**(1/3)  

    # voorspelling voor 2026
    pop_2026 = pop_2024 * (g)**2
    population_2026_dict[icao] = pop_2026

gdp_2026_dict = {}
for icao in gdp_2021_dict:
    gdp_2021 = gdp_2021_dict[icao]
    gdp_2024 = gdp_2024_dict[icao]

    # Grow factor, in 3 years de difference is grow
    g = (gdp_2024 / gdp_2021)**(1/3)

    # 2026 is amount of 2024 times the grow factor to the power of the difference in years
    gdp_2026 = gdp_2024 * (g)**2
    gdp_2026_dict[icao] = gdp_2026

# print(f'Population in 2026')
# print(gdp_2026_dict)

# print(f'GDP in 2026')
# print(population_2026_dict)

# Demand prediction 2026
D_pred_26 = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred_26[i, j] = 0  # geen vraag van een stad naar zichzelf
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product_26 = population_2026_dict[i_icao] * population_2026_dict[j_icao]
        gdp_product_26 = gdp_2026_dict[i_icao] * gdp_2026_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred_26[i, j] = k1 * (pop_product_26**b1) * (gdp_product_26**b2) * ((fd_ij)**(beta3))

# print("Voorspelde demand 2026 (gravity model):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D_pred_26[i,j]:.0f}" for j in range(n)))

# Voorspelde ln(D) met het model
df['lnD_pred'] = model.predict(X)

# Omzetten naar originele schaal
df['D_pred'] = np.exp(df['lnD_pred'])
df['D_actual'] = np.exp(df['lnD'])

# # Plot: Werkelijk vs Voorspeld D
# plt.figure(figsize=(8,6))
# plt.scatter(df['D_actual'], df['D_pred'], color='blue', alpha=0.6, label='Data points')
# plt.plot([0, df['D_actual'].max()], [0, df['D_actual'].max()],
#          color='red', lw=2, label='Regression line')
# plt.xlabel('Given demand')
# plt.ylabel('Predicted demand')
# plt.title('Given vs. predicted demand 2021')
# plt.legend()
# plt.grid(True)
# plt.show()


### Implementatie van mathematical model ###

from gurobipy import *

# Inladen aircraft
# --- Workbook openen ---
wb2 = openpyxl.load_workbook("AircraftData.xlsx")
sheet2 = wb2.active   # neem het eerste werkblad

# Lees de header (aircraft names)
aircraft_names = [cell.value for cell in sheet2[1][1:]]  # rij 1, vanaf kolom B

data = {}

# Loop door rijen en vul dictionary
for row in sheet2.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        continue
    
    parameter_name = row[0]
    values = row[1:]
    
    data[parameter_name] = values

# DataFrame bouwen (transpose zodat aircraft types rijen worden)
df_aircraft = pd.DataFrame(data, index=aircraft_names)
df_aircraft = df_aircraft.drop(columns=["Aircraft Characteristics"])

print(df_aircraft)


# Data - Sets
N = range(len(airports))                    # Set of airports; i, j in N
K = range(len(df_aircraft))                 # Set of aircrafts; k in K
ac = len(df_aircraft)                        # Total amount of aircrafts

# Data - parameters
q = np.zeros((n, n))                        # Demand between orgin i and destination j
for i in N:
    for j in N:
        q[i, j] = D_pred_26[i, j]

d = np.zeros((n, n))                        # Distance between i and j
for i in N:
    for j in N:
        d[i, j] = dij[i, j]


y = np.zeros((n, n))
for i in N:                                 #YIELD
    for j in N:
        if d[i, j] > 0:
            y[i, j] = 5.9 * d[i, j]**(-0.76) + 0.043
        else:
            y[i, j] = 0   


s = np.zeros(ac)
for k in K:
    s[k] = df_aircraft['Seats'][k]

v = np.zeros(ac)
for k in K:
    v[k] = df_aircraft['Speed [km/h]'][k]

t = np.zeros(ac)
for k in K:
    t[k] = df_aircraft['Average TAT [mins]'][k]

ra = np.zeros(ac)
for k in K:
    ra[k] = df_aircraft['Maximum range [km]'][k]

RAC = np.zeros(ac)
for k in K:
    RAC[k] = df_aircraft['Runway required [m]'][k]

RAP = np.zeros(n)
for j in N:
    RAP[j] = runways[j]

cl = np.zeros(ac)
for k in K:
    cl[k] = df_aircraft['Weekly lease cost [€]'][k]
    print(cl[k])

C = np.zeros(ac)
for k in K:
    C[k] = df_aircraft['Fixed operating cost C_X [€]'][k]

CT = np.zeros(ac)
for k in K:
    CT[k] = df_aircraft['Time cost parameter C_T [€/hr]'][k]

C_Tij = np.zeros((n, n, ac))
for k in K:
    C_Tij[ :, :,k] = CT[k] * (d/ v[k])
    print(C_Tij)

CF = np.zeros(ac)
for k in K:
    CF[k] = df_aircraft['Fuel cost parameter C_F'][k]

C_Fij = np.zeros((n, n, ac))
for k in K:
    C_Fij[:, :, k] = ((CF[k] * 1.42) / 1.5) * d

# for k in K: -->waarom deze regels code
#     (f'For Aircraf {k} fuel cost is {C_Fij[k]}')

Ck_ij = np.zeros((n, n, ac))
for k in K:
    for i in N:
        for j in N:
            Ck_ij[i, j, k] = C[k] + C_Tij[i, j, k] + C_Fij[i, j, k]
            

a = {}
for i in N:
    for j in N:
        for k in K:
            if d[i,j] <= ra[k]:
                a[i,j,k] = 10000
            else:
                a[i,j,k] = 0

g = np.ones(n)
g[hub_index] = 0 

LF = 0.75                                                              # Loadfactor, given

BT = 70                                                                # Available hours in a week per airplaine

TS = np.zeros(n)                                                       # Amount of available time slots
for j in N:
    TS[j] = slots[j]

TAT = np.zeros(ac)
for k in K:
    TAT[k] = df_aircraft['Average TAT [mins]'][k]/60


#hier moeten de laaste paar parameters nog komen. Ik ga alvast verder met het model

from gurobipy import Model, GRB, quicksum

def main():
    model = Model("Model_1B")
    model.write("network_fleet_development.lp")
    model.setParam('TimeLimit', 120)

    x = model.addVars(N, N, name="x", vtype=GRB.INTEGER, lb=0)
    w = model.addVars(N, N, name="w", vtype=GRB.INTEGER, lb=0)
    z = model.addVars(N, N, K, name="z", vtype=GRB.INTEGER, lb=0)
    AC = model.addVars(K, name="AC", vtype=GRB.INTEGER, lb=0)

    #objective:
    
    Objective_1B = quicksum((y[i, j] * d[i, j] * (x[i, j] + w[i, j])) for i in N for j in N) - quicksum(Ck_ij[i, j, k] * z[i, j, k] for k in K for i in N for j in N) - quicksum(cl[k] * AC[k] for k in K)
    

    model.setObjective(Objective_1B, GRB.MAXIMIZE)

    # Constraints:
    # model.addConstr(quicksum(x[i, j] >= 1 for i in N for j in N), name="at_least_one_flight")
    # model.addConstr(quicksum(w[i, j] >= 1 for i in N for j in N), name="at_least_one_flight")
    # model.addConstr(quicksum(z[i, j, k] >= 1 for i in N for j in N for k in K), name="at_least_one_flight")
    # model.addConstr(quicksum(AC[k] >= 1 for k in K), name="at_least_one_aircraft")



    #passengers smaller than demand
    for i in N:
        for j in N:
            model.addConstr(x[i, j] + w[i, j] <= q[i, j], name=f"demand_limit_{i}_{j}")

    for i in N:
        for j in N:
            model.addConstr(w[i,j]<= q[i,j] * g[i] * g[j],name=f"transfer")

    # #nieuwe constrain die dwingt dat er altijd via de hub wordt gevlogen (martijn)
    # for i in N:
    #     for j in N:
    #         for k in K:
    #             model.addConstr(z[i, j, k] * g[i] * g[j] == 0, name=f"via_hub_{i}_{j}_{k}")

    # hub constr die dwingt dat er altijd via de hub wordt gevlogen (do)
    for i in N:
        for j in N:
            if i != hub_index and j != hub_index:
                for k in K:
                    model.addConstr(z[i, j, k] == 0, name=f"hub_constr_{i}_{j}_{k}")


    #balance incomming and outgoing
    for i in N:
        for k in K:
            lhs = quicksum(z[i, j, k]  for j in N)
            rhs = quicksum(z[j, i, k]  for j in N)
            model.addConstr(lhs == rhs, name=f"return_constrain_{i}_{k}")
    
    # passengers smaller than amount of seats 
    for i in N:
        for j in N:
            lhs1 = x[i,j] + quicksum((w[i,m]*(1 - g[j])) for m in N) + quicksum((w[m,j]*(1 - g[i])) for m in N)
            rhs1 = quicksum((z[i, j, k]  * s[k] * LF) for k in K)
            model.addConstr(lhs1 <= rhs1, name=f"cap_constraint_{i}_{j}")

# # passengers smaller than amount of seats met toevoeging martijn
#     for i in N:
#         for j in N:
#             lhs1 = (x[i, j] + quicksum(w[i, m] * (1 - g[j]) for m in N if m != i and m != j) + quicksum(w[m, j] * (1 - g[i]) for m in N if m != i and m != j))
#             rhs1 = quicksum((z[i, j, k]  * s[k] * LF) for k in K)
#             model.addConstr(lhs1 <= rhs1, name=f"cap_constraint_{i}_{j}")


    # #duration of flights basic
    # for k in K:
    #    lhs2 = quicksum( (((d[i, j] / v[k]) + TAT[k]) * z[i, j, k] ) for i in N for j in N)
    #    rhs2 = (BT * AC[k])
    #    model.addConstr(lhs2 <= rhs2)

    #    #duration poging 200 --> met haakjes om de totale som van de tijd * z_ijk geeft alles 0, maar moet wel zo denk ik
    # for k in K:
    #     lhs3 = quicksum(((d[i,j]/ v[k]) + TAT[k] + (TAT[k] * 0.5 * (1 - (g[i] * g[j] )))) * z[i, j, k]  for i in N for j in N) 
    #     rhs3 = (BT * AC[k]) 
    #     model.addConstr(lhs3 <= rhs3, name=f"duration_constrain_{k}")

           #duration poging 300 --> geen haakjes om de totale som van de tijd * z_ijk, dit klopt eigenlijk niet, maar zo krijg je wel een oplossing
    for k in K:
        lhs3 = quicksum((d[i,j]/ v[k]) + TAT[k] + (TAT[k] * 0.5 * (1 - (g[i] * g[j] ))) * z[i, j, k]  for i in N for j in N) 
        rhs3 = (BT * AC[k]) 
        model.addConstr(lhs3 <= rhs3, name=f"duration_constrain_{k}")


    # range constrains
    for k in K:
        for i in N:
            for j in N:
                model.addConstr(z[i, j, k]  <= a[i, j, k], name=f"reach_{i}_{j}_{k}")

    #runway constraints
    for k in K:
        for i in N:
            for j in N:
                model.addConstr(RAC[k] * z[i, j, k] <= RAP[i]* z[i, j, k] , name=f'Runway_dep_{i}_{j}_{k}') # heb z_ijk toegevoegd zodat er wel iets wordt gedaan met dat er een vliegtuig heen gaat

    for k in K:
        for i in N:
            for j in N:
                model.addConstr(RAC[k] * z[i, j, k]  <= RAP[j] * z[i, j, k] , name=f'Runway_arr_{i}_{j}_{k}') # heb z_ijk toegevoegd zodat er wel iets wordt gedaan met dat er een vliegtuig heen gaat

    #timeslot constraints
    for j in N:
        model.addConstr(quicksum(z[i, j, k]  for i in N for k in K) <= TS[j], name=f'Time_slots_{j}')
    
  


    model.optimize()
    model.write("Model_1B.sol")
    model.write("Model_1B.lp")
    if model.status == GRB.OPTIMAL or model.status == GRB.TIME_LIMIT:
        totale_winst = model.ObjVal  
        print(f"Totaal winst: €{totale_winst:.2f}")

        print("\nAantal benodigde vliegtuigen per type:")
        for k in K:
            print(f"{df_aircraft.index[k]}: {AC[k].X:.0f}")
            
        totaal_passagiers = sum(x[i, j].X for i in N for j in N)
        print(f"\nTotaal aantal passagiers die reizen van of naar de hub: {totaal_passagiers:.0f}")

        transfer_passagiers = sum(w[i, j].X for i in N for j in N)
        print(f'\nAantal passagiers die transfer hebben op de hub {transfer_passagiers}')

        print("\nGevlogen routes (z[i,j,k] > 0):")
       
        for i in N:
            for j in N:
                for k in K:
                    if z[i, j, k].X > 0.5:   # threshold om integer rounding te vermijden
                        print(f"{airports[i]} → {airports[j]} met {df_aircraft.index[k]} "
                            f"aantal vluchten: {z[i, j, k].X:.0f}")
                        
        print("\nVliegtijd per individueel vliegtuig:")
        for k in K:
            total_hours_k = sum(
                ((d[i, j] / v[k]) + TAT[k] + (TAT[k] * 0.5 * (1 -g[j]))) * z[i, j, k].X
                for i in N for j in N)

            num_aircraft = AC[k].X

            if num_aircraft > 0:
                hours_per_aircraft = total_hours_k / num_aircraft
                print(f"Type {k}: {hours_per_aircraft:.2f} uur per vliegtuig")
            else:
                print(f"Type {k}: geen vliegtuigen gebruikt")


        print("Matrix unmet demand = demand – (w + x):\n")
        unmet = np.zeros(n)
        for j in N:
            sum_z = sum(z[i, j, k].X for i in N for k in K)  # waarde van z na optimalisatie
            unmet[j] = TS[j] - sum_z
            print(f"{airports[j]}: {unmet[j]:.0f}")



        print("\nMatrix x[i,j]:")
        for i in N:
            row = ""
            for j in N:
                row += f"{x[i,j].X:8.1f} "
            print(row)

        print("\nMatrix w[i,j]:")
        for i in N:
            row = ""
            for j in N:
                row += f"{w[i,j].X:8.1f} "
            print(row)


        # Header
        print("\t" + "\t".join(airports))

        for i in N:
            row_values = []
            for j in N:
                demand = D_pred_26[i, j]
                flow = w[i, j].X + x[i, j].X    # totaal vervoerd
                unmet = demand - flow
                row_values.append(f"{unmet:.0f}")
            print(f"{airports[i]}\t" + "\t".join(row_values))

        for k in K:
            print(f"\nAircraft type {k}:")
            for i in N:
             for j in N:
                    flight_distance = d[i,j]
                    speed = v[k]
                    tat = TAT[k]
                    g_factor = g[j]

                    # Bereken de drie onderdelen van de term
                    term1 = flight_distance / speed
                    term2 = tat
                    term3 = tat * 0.5 * (1 - g_factor)

                    # Totale duration_term
                    duration_term = term1 + term2 + term3

                    # Print alleen relevante info
                    if abs(duration_term) < 1e-6:
                        print(f"Flight {i}->{j} duration_term = 0")
                        print(f"  term1 (d/v) = {term1}")
                        print(f"  term2 (TAT) = {term2}")
                        print(f"  term3 (TAT*0.5*(1-g)) = {term3}")
                    elif z[i, j, k].X > 1e-6:
                        print(f"Flight {i}->{j} contributes: duration_term = {duration_term:.2f}, z = {z[i, j, k].X}")

        print("Ck_ij shape:", Ck_ij.shape)
        print("z vars:", len(z))
        i,j,k = 0,1,2
        print("Ck_ij[k,i,j] =", Ck_ij[k,i,j])
        print("z[i,j,k] =", z[i, j, k].X if model.status == GRB.OPTIMAL else "NA")



    else:
        print("No optimal solution found")

main()
   