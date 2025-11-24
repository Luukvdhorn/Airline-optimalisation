from openpyxl import *
from time import *
import numpy as np
import gurobipy as gp
from gurobipy import GRB
import math
import matplotlib.pyplot as plt
import pandas as pd
import os as os

class Node:
    def __init__(self, name, country, pop2021, pop2024, gdp2021, gdp2024):
        self.Name = name
        self.Country = country
        self.Pop2021 = pop2021
        self.Pop2024 = pop2024
        self.GDP2021 = gdp2021
        self.GDP2024 = gdp2024
        
        self.InLinks = []
        self.OutLinks = []

    def addInLink(self, node):
        self.InLinks.append(node)

    def addOutLink(self, node):
        self.OutLinks.append(node)


class Arc:
    def __init__(self, origin, destination, gdp_i, gdp_j):
        self.From = origin
        self.To = destination
        self.GDP_i = gdp_i
        self.GDP_j = gdp_j


# =====================================================================
#   Construct graph from your pop.xlsx + GDP structure
# =====================================================================

def construct_graph_from_population_file():
    import pandas as pd
    from openpyxl import load_workbook

    # Load Excel
    wb = load_workbook("pop.xlsx", read_only=True)
    sheet = wb.active
    rows = tuple(sheet.iter_rows(values_only=True))

    # Find header row
    header_index = None
    for i, row in enumerate(rows):
        if row and "City" in row:
            header_index = i
            break

    if header_index is None:
        raise ValueError("Header 'City' not found!")

    # Extract population rows
    pop_rows = []
    for r in rows[header_index+1:]:
        if r[0] is None:
            break
        pop_rows.append(r[:3])  # city, pop21, pop24

    df_pop = pd.DataFrame(pop_rows, columns=["City", "Pop2021", "Pop2024"])

    # Extract GDP rows
    gdp_rows = []
    for r in rows[header_index+1:]:
        if len(r) < 7 or r[4] is None:
            break
        gdp_rows.append(r[4:7])  # country, gdp21, gdp24

    df_gdp = pd.DataFrame(gdp_rows, columns=["Country", "GDP2021", "GDP2024"])

    # Combine rows: assumption = same order (as in your file)
    df = pd.concat([df_pop, df_gdp], axis=1)

    # Build graph objects
    Nodes = []
    Arcs = []

    # Nodes: one per city
    for i, row in df.iterrows():
        node = Node(
            name=row["City"],
            country=row["Country"],
            pop2021=row["Pop2021"],
            pop2024=row["Pop2024"],
            gdp2021=row["GDP2021"],
            gdp2024=row["GDP2024"]
        )
        Nodes.append(node)

    # Arcs: all combinations i→j with GDP values
    for i in range(len(Nodes)):
        for j in range(len(Nodes)):
            if i != j:
                arc = Arc(
                    origin=i,
                    destination=j,
                    gdp_i=Nodes[i].GDP2024,
                    gdp_j=Nodes[j].GDP2024
                )
                Arcs.append(arc)
                Nodes[i].addOutLink(j)
                Nodes[j].addInLink(i)

    return Nodes, Arcs
