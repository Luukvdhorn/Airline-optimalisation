from gurobipy import * 
from openpyxl import * 
import openpyxl
from time import *
import numpy as np
import math
import matplotlib.pyplot as plt
import pandas as pd
import statsmodels.api as sm

#IMPORT ALL THE EXCEL FILES FOR ASSIGNMENT 1A

wb = load_workbook("pop.xlsx", data_only=True)
ws = wb["General"]

wb_codes = load_workbook("Airport_names.xlsx", data_only=True)
ws_codes = wb_codes.active

city_to_icao = {}
country_to_icao = {}

row = 1
while True:                                                             #Give everything a good name and couple them
    city = ws_codes.cell(row=row, column=1).value
    country = ws_codes.cell(row=row, column=2).value
    icao = ws_codes.cell(row=row, column=3).value

    if city is None:
        break

    city_to_icao[city] = icao
    country_to_icao[country] = icao

    row += 1

population_2021_dict = {}
population_2024_dict = {}
gdp_2021_dict = {}
gdp_2024_dict = {}

row = 4

#GET THE RIGHT INFORMATION FROM THE EXCELFILES

while True:
    city = ws.cell(row=row, column=1).value        # kolom A
    pop2021 = ws.cell(row=row, column=2).value     # kolom B
    pop2024 = ws.cell(row=row, column=3).value     # kolom C
    gdp2021 = ws.cell(row=row, column=6).value     # kolom F
    gdp2024 = ws.cell(row=row, column=7).value     # kolom G

    if city is None:
        break

    if city in city_to_icao:
        icao = city_to_icao[city]
        population_2021_dict[icao] = pop2021
        population_2024_dict[icao] = pop2024
        gdp_2021_dict[icao] = gdp2021
        gdp_2024_dict[icao] = gdp2024

    row += 1

wb = load_workbook("DemandGroup40.xlsx", data_only=True)
ws = wb.active

#INFORMATION FOR ASSIGNMENT 1B
icao_row = 5    
lat_row = 6    
lon_row = 7    
start_col = 3
runway_row = 8
slots_row  = 9
runways = []
slots = []
hub_index = 2                                   # Amsterdam

RE = 6371.0                                     # Radius Earth
f = 1.42                                        # Fuel cost constant 

airports = []
latitudes = []
longitudes = []

col = start_col
while True:
    icao = ws.cell(row=icao_row, column=col).value
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    runway = ws.cell(row=runway_row, column=col).value
    slot = ws.cell(row=slots_row, column=col).value
    if icao is None:
        break
    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    runways.append(float(runway))
    if slot == '-':
        slots.append(float('inf'))
    else:
        slots.append(float(slot))
    col += 1

#DISTANCE FORMULA

def distance(phi_i, lam_i, phi_j, lam_j):
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 +np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))


n = len(airports)
dij = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        dij[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])


demand_start_row = icao_row + 8 
demand_start_col = start_col 


#BUILDING UP D (IJ)

n = len(airports)
D = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        cell = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(cell) if cell is not None else 0
        except:
            D[i, j] = 0

n = len(airports)
rows = []

for i in range(n):        
    for j in range(n):
        if i == j:
            continue                            # If orgin and destionation is the same, go on

        Dij = D[i, j]
        if Dij <= 0:
            continue                            # It there is no demand, go on

        i_icao = airports[i]                    #Collectiong ICAO code
        j_icao = airports[j]

        row = {
            "D": Dij,
            "pop": population_2021_dict[i_icao] * population_2021_dict[j_icao],
            "gdp": gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao],
            "fd": f * dij[i, j]
        }
        rows.append(row)

df = pd.DataFrame(rows)

#LINEAR REGRESSION

df['lnD']    = np.log(df['D'])                  
df['ln_pop'] = np.log(df['pop'])
df['ln_gdp'] = np.log(df['gdp'])
df['ln_fd']  = np.log(df['fd'])


X = df[['ln_pop', 'ln_gdp', 'ln_fd']]           # Independent variables
X = sm.add_constant(X)                          # This creates ln(k)
y = df['lnD']                                   # Dependent variable

model = sm.OLS(y, X).fit()                      # Ordinary least squares
#print(model.summary())


a = model.params['const']            
b1 = model.params['ln_pop']         
b2 = model.params['ln_gdp']

beta3 = model.params['ln_fd']        
b3 = -beta3                          

k1 = np.exp(a)                       


# print(f"k1  = {k1}")                          #printing the k, b1, b2,b3 values
# print(f"b1 = {b1:.4f}")
# print(f"b2 = {b2:.4f}")
# print(f"b3 = {b3:.4f}")


#CREATE PREDICITION 2021

D_pred = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred[i, j] = 0  
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product = population_2021_dict[i_icao] * population_2021_dict[j_icao]
        gdp_product = gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred[i, j] = k1 * (pop_product**b1) * (gdp_product**b2) * ((fd_ij)**(beta3))


# DIFFERENCE MATRIX
D_diff = D_pred - D  


# POP EN GDP 2021 AND 2024 FOR GROWFACTORS
population_2026_dict = {}
pop_g_list = []                                         #Average pop g list

for icao in population_2021_dict:
    pop_2021 = population_2021_dict[icao]
    pop_2024 = population_2024_dict[icao]

    g = (pop_2024 / pop_2021)**(1/3)                   #Annual growth
    pop_g_list.append(g)
    pop_2026 = pop_2024 * (g)**2
    population_2026_dict[icao] = pop_2026              #Prediction 2026




gdp_2026_dict = {}
gdp_g_list = []

for icao in gdp_2021_dict:
    gdp_2021 = gdp_2021_dict[icao]
    gdp_2024 = gdp_2024_dict[icao]

    g = (gdp_2024 / gdp_2021)**(1/3)
    gdp_g_list.append(g)
    
    gdp_2026 = gdp_2024 * (g)**2
    gdp_2026_dict[icao] = gdp_2026


avg_gdp_g = sum(gdp_g_list) / len(gdp_g_list)
#print(f"Gemiddelde jaarlijkse groeifactor GDP: {avg_gdp_g:.4f}")

avg_pop_g = sum(pop_g_list) / len(pop_g_list)
#print(f"Gemiddelde jaarlijkse groeifactor populatie: {avg_pop_g:.4f}")

#PREDICTION 2026
D_pred_26 = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred_26[i, j] = 0  
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product_26 = population_2026_dict[i_icao] * population_2026_dict[j_icao]
        gdp_product_26 = gdp_2026_dict[i_icao] * gdp_2026_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred_26[i, j] = k1 * (pop_product_26**b1) * (gdp_product_26**b2) * ((fd_ij)**(beta3))


D = np.array(D_pred_26)

df['lnD_pred'] = model.predict(X)

df['D_pred'] = np.exp(df['lnD_pred'])
df['D_actual'] = np.exp(df['lnD'])

x = df['D_actual']
y = df['D_pred']



#PLOTS AND PRINTSTATEMENTS

fig, ax = plt.subplots(figsize=(14, 12))
im = ax.imshow(D, cmap="Greens")

ax.set_xticks(np.arange(len(airports)))
ax.set_yticks(np.arange(len(airports)))
ax.set_xticklabels(airports)
ax.set_yticklabels(airports)

ax.tick_params(top=False, bottom=True, labeltop=False, labelbottom=True)

plt.setp(
    ax.get_xticklabels(),
    rotation=90,
    ha="center",
    va="top"
)


for i in range(len(airports)):
    for j in range(len(airports)):
        if D[i, j] > 0: 
                ax.text(
                j, i, f"{D[i, j]:.0f}",
                ha="center", va="center",
                fontsize=8,
                color="black")

ax.set_title("Estimated demand 2026")

plt.tight_layout()
plt.show()



# Plot: Werkelijk vs Voorspeld D
plt.figure(figsize=(8,6))
plt.scatter(df['D_actual'], df['D_pred'], color='blue', alpha=0.6, label='Data points')
plt.plot([0, df['D_actual'].max()], [0, df['D_actual'].max()],
         color='red', lw=2, label='Regression line')
plt.xlabel('Given demand')
plt.ylabel('Predicted demand')
plt.title('Given vs. predicted demand 2021')
plt.legend()
plt.grid(True)
plt.show()

# Richtingscoëfficiënt (slope) berekenen
slope, intercept = np.polyfit(x, y, 1)
print(f"Richtingscoëfficiënt (slope): {slope:.4f}")












