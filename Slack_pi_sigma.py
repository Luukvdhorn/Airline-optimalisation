

### GESLASHEDDD OMDAT DIT CODE IS VOOR SIGMA EN PI


# from gurobipy import *
# from openpyxl import load_workbook
# import pandas as pd
# import numpy as np

# # =========================
# # DATA INLADEN
# # =========================

# wb_data = load_workbook("data_ex2.xlsx", data_only=True)

# df_flights = pd.DataFrame(wb_data["Flights"].values)
# df_flights.columns = df_flights.iloc[0]
# df_flights = df_flights.iloc[1:].reset_index(drop=True)

# df_itineraries = pd.DataFrame(wb_data["Itineraries"].values)
# df_itineraries.columns = df_itineraries.iloc[0]
# df_itineraries = df_itineraries.iloc[1:].reset_index(drop=True)

# # =========================
# # SETS
# # =========================

# L = range(len(df_flights))                 # flights
# P_real = range(len(df_itineraries))        # echte itineraries
# p_fict = len(df_itineraries)               # fictitious itinerary = 0 in slides
# P = list(P_real) + [p_fict]

# # =========================
# # PARAMETERS
# # =========================

# # fare
# fare = {}
# for p in P_real:
#     fare[p] = float(df_itineraries.loc[p, "Price [EUR]"])
# fare[p_fict] = 0.0                         # slide 24

# # demand
# D = {}
# for p in P_real:
#     D[p] = float(df_itineraries.loc[p, "Demand"])
# D[p_fict] = 100000    # big M

# # capacity
# CAP = {}
# for l in L:
#     CAP[l] = float(df_flights.loc[l, "Capacity"])

# # =========================
# # DELTA_lp
# # =========================

# flight_code_to_id = {df_flights.loc[l, 'Flight No.']: l for l in L}
# delta_lp = {}

# for p, row in df_itineraries.iterrows():

#     if pd.notna(row['Flight 1']):
#         l = flight_code_to_id[row['Flight 1']]
#         delta_lp[(l, p)] = 1

#     if pd.notna(row['Flight 2']):
#         l = flight_code_to_id[row['Flight 2']]
#         delta_lp[(l, p)] = 1

# # fictitious itinerary gebruikt geen vluchten
# for l in L:
#     delta_lp[(l, p_fict)] = 0

# # =========================
# # KEYPATH PARAMETER Q_l
# # =========================

# Q = {}
# for l in L:
#     Q[l] = sum(delta_lp.get((l, p), 0) * D[p] for p in P_real)

# # =========================
# # b_pr – EXACT SLIDE 24
# # =========================

# b_pr = {}

# # b_pp = 1
# for p in P_real:
#     b_pr[(p, p)] = 1.0

# # b_0p = 1
# for p in P_real:
#     b_pr[(p, p_fict)] = 1.0

# # =========================
# # MODEL: INITIAL RMP
# # =========================

# def main():

#     model = Model("RMP_slide_24")
#     model.setParam("TimeLimit", 300)

#     # -------------------------
#     # DECISION VARIABLES
#     # Alleen t_{0p}
#     # -------------------------

#     t = {}
#     for p in P_real:
#         t[p, p_fict] = model.addVar(lb=0, vtype=GRB.INTEGER, name=f"t_{p}_{p_fict}")

#     # -------------------------
#     # OBJECTIVE FUNCTION
#     # -------------------------
#     # min sum_p fare_p * t_0p

#     model.setObjective(quicksum(fare[p] * t[p, p_fict] for p in P_real), GRB.MINIMIZE)

#     # -------------------------
#     # CAPACITY CONSTRAINTS
#     # -------------------------

#     for l in L:
#         model.addConstr(
#             quicksum(delta_lp.get((l, p), 0) * t[p, p_fict] for p in P_real)
#             >= Q[l] - CAP[l],
#             name=f"cap_{l}"
#         )


#     # -------------------------
#     # DEMAND CONSTRAINTS
#     # -------------------------

#     for p in P_real:
#         model.addConstr(t[p, p_fict] <= D[p],
#             name=f"demand_{p}")

#     # -------------------------
#     # SOLVE
#     # -------------------------

#     model.optimize()
#     model.write("RMP_slide_24.lp")
#     # -------------------------
#     print(p_fict)
    
#     print(P_real)

#     print(P)
#     # -------------------------
#     # OUTPUT
#     # -------------------------

#     if model.status == GRB.OPTIMAL:
#         print(f"\nObjective value (lost revenue): {model.objVal:.2f}\n")
#         print("Spilled passengers (t_0p):")
#         for p in P_real:
#             if t[p, p_fict].x > 1e-6:
#                 print(f"t[0,{p}] = {t[p, p_fict].x:.2f}")
#     else:
#         print("No feasible solution found.")

# main()
from gurobipy import *
from openpyxl import load_workbook
import pandas as pd
import numpy as np

# =========================
# DATA INLADEN
# =========================

wb_data = load_workbook("data_ex2.xlsx", data_only=True)

df_flights = pd.DataFrame(wb_data["Flights"].values)
df_flights.columns = df_flights.iloc[0]
df_flights = df_flights.iloc[1:].reset_index(drop=True)

df_itineraries = pd.DataFrame(wb_data["Itineraries"].values)
df_itineraries.columns = df_itineraries.iloc[0]
df_itineraries = df_itineraries.iloc[1:].reset_index(drop=True)

df_recapture = pd.DataFrame(wb_data["Recapture"].values)
df_recapture.columns = df_recapture.iloc[0]
df_recapture = df_recapture.iloc[1:].reset_index(drop=True)

# =========================
# SETS
# =========================

L = range(len(df_flights))                 # flights
P_real = range(len(df_itineraries))        # echte itineraries
p_fict = len(df_itineraries)               # fictitious itinerary index
P = list(P_real) + [p_fict]

# =========================
# PARAMETERS
# =========================

# Fare
fare = {}
for p in P_real:
    fare[p] = float(df_itineraries.loc[p, "Price [EUR]"])
fare[p_fict] = 0.0

# Demand
D = {}
for p in P_real:
    D[p] = float(df_itineraries.loc[p, "Demand"])
D[p_fict] = sum(D[p] for p in P_real)  # groot genoeg

# Capacity
CAP = {}
for l in L:
    CAP[l] = float(df_flights.loc[l, "Capacity"])

# =========================
# DELTA_lp
# =========================

flight_code_to_id = {df_flights.loc[l, 'Flight No.']: l for l in L}
delta_lp = {}

for p, row in df_itineraries.iterrows():

    code1 = row['Flight 1']
    if pd.notna(code1):
        l = flight_code_to_id[code1]
        delta_lp[(l, p)] = 1

    code2 = row['Flight 2']
    if pd.notna(code2):
        l = flight_code_to_id[code2]
        delta_lp[(l, p)] = 1

# fictitious itinerary gebruikt geen vluchten
for l in L:
    delta_lp[(l, p_fict)] = 0

# =========================
# KEYPATH PARAMETERS
# =========================

Q = {}
for l in L:
    Q[l] = sum(delta_lp.get((l, p), 0) * D[p] for p in P_real)

# =========================
# b_pr (ALLEEN self + fictitious)
# =========================

b_pr = {}

for p in P_real:
    b_pr[(p, p)] = 1.0       # self
    b_pr[(p, p_fict)] = 1.0 # fictitious

# =========================
# MODEL: INITIAL RMP
# =========================

def main():

    model = Model("Passenger_Mix_Flow_RMP")
    model.setParam("TimeLimit", 300)

    # -------------------------
    # DECISION VARIABLES
    # Alleen t_{0p}
    # -------------------------

    t = {}
    for p in P_real:
        t[p, p_fict] = model.addVar(
            lb=0,
            vtype=GRB.CONTINUOUS,
            name=f"t_{p}_{p_fict}"
        )

    model.update()

    # -------------------------
    # OBJECTIVE FUNCTION
    # -------------------------

    model.setObjective(
        quicksum(fare[p] * t[p, p_fict] for p in P_real),
        GRB.MINIMIZE
    )

    # -------------------------
    # CAPACITY CONSTRAINTS
    # -------------------------

    cap_constr = {}
    for l in L:
        cap_constr[l] = model.addConstr(
            quicksum(delta_lp.get((l, p), 0) * t[p, p_fict] for p in P_real)
            >= Q[l] - CAP[l],
            name=f"cap_{l}"
        )

    # -------------------------
    # DEMAND CONSTRAINTS
    # -------------------------

    demand_constr = {}
    for p in P_real:
        demand_constr[p] = model.addConstr(t[p, p_fict] <= D[p],
            name=f"demand_{p}"
        )

    # -------------------------
    # SOLVE
    # -------------------------

    model.optimize()
    model.write("RMP_iteration_0.lp")

    # =========================
# DUAL VARIABLES
# =========================

    pi = {l: cap_constr[l].Pi for l in L}
    sigma = {p: demand_constr[p].Pi for p in P_real}

    def sum_pi_over_itinerary(p):
        return sum(pi[l] for l in L if delta_lp.get((l, p), 0) == 1)




    # -------------------------
    # OUTPUT
    # -------------------------

    if model.status in [GRB.OPTIMAL, GRB.TIME_LIMIT]:
        print(f"\nObjective value (lost revenue): {model.objVal:.2f}\n")

        print("Spilled passengers to fictitious itinerary:")
        for p in P_real:
            if t[p, p_fict].x > 1e-6:
                print(f"  t[0,{p}] = {t[p, p_fict].x:.2f}")

        print("\nDual variables:")
        for l in L:
            if cap_constr[l].Pi != 0:
                print(f"pi[{l}] = {cap_constr[l].Pi:.2f}")
        
        for p in P_real:
            if demand_constr[p].Pi != 0:
                print(f"  sigma[{p}] = {demand_constr[p].Pi:.2f}")
        print("\nPricing problem – slack (reduced costs):")

        for _, row in df_recapture.iterrows():

            p = int(row["From Itinerary"])
            r = int(row["To Itinerary"])
            b_rp = float(row["Recapture Rate"])

            # modified fares (slide 32)
            mod_fare_p = fare[p] - sum_pi_over_itinerary(p)
            mod_fare_r = fare[r] - sum_pi_over_itinerary(r)

            # reduced cost / slack
            slack = mod_fare_p - b_rp * mod_fare_r - sigma[p]
            if slack < 0:
                print(f"c'[{r},{p}] = {slack:.2f}")

                if slack < 0:
                    print(f"  -> column t[{r},{p}] PRICES OUT (ADD)")

    else:
        print("No feasible solution found.")

main()
