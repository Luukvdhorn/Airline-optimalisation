from gurobipy import * 
from openpyxl import * 
import openpyxl
from time import *
import numpy as np
import math
import matplotlib.pyplot as plt
import pandas as pd
import statsmodels.api as sm

wb = load_workbook("pop.xlsx", data_only=True)
ws = wb["General"]

population_2021 = []
gdp_2021 = []

population_2024 = []
gdp_2024 = []

row = 4  
while True:
    city = ws.cell(row=row, column=1).value          # kolom A
    pop2021 = ws.cell(row=row, column=2).value       # kolom B

    if city is None:
        break

    population_2021.append((city, pop2021))
    row += 1

row=4

while True:
    city = ws.cell(row=row, column=1).value          # kolom A
    pop2024 = ws.cell(row=row, column=3).value       # kolom C

    if city is None:
        break

    population_2024.append((city, pop2024))
    row += 1

row = 4  

while True:
    country = ws.cell(row=row, column=5).value       # kolom E
    gdp2021 = ws.cell(row=row, column=6).value       # kolom F

    if country is None:
        break  
    gdp_2021.append((country, gdp2021))
    row += 1

row = 4

while True:
    country = ws.cell(row=row, column=5).value       # kolom E
    gdp2024 = ws.cell(row=row, column=7).value       # kolom G

    if country is None:
        break  
    gdp_2024.append((country, gdp2024))
    row += 1


wb_codes = load_workbook("Airport_names.xlsx", data_only=True)
ws_codes = wb_codes.active

city_to_icao = {}
country_to_icao = {}

row = 1
while True:
    city = ws_codes.cell(row=row, column=1).value
    country = ws_codes.cell(row=row, column=2).value
    icao = ws_codes.cell(row=row, column=3).value

    if city is None:
        break

    city_to_icao[city] = icao
    country_to_icao[country] = icao

    row += 1

population_2021_dict = {}
population_2024_dict = {}
gdp_2021_dict = {}
gdp_2024_dict = {}

row = 4
while True:
    city = ws.cell(row=row, column=1).value        # kolom A
    pop2021 = ws.cell(row=row, column=2).value     # kolom B
    pop2024 = ws.cell(row=row, column=3).value     # kolom C
    gdp2021 = ws.cell(row=row, column=6).value     # kolom F
    gdp2024 = ws.cell(row=row, column=7).value     # kolom G

    if city is None:
        break

    if city in city_to_icao:
        icao = city_to_icao[city]
        population_2021_dict[icao] = pop2021
        population_2024_dict[icao] = pop2024
        gdp_2021_dict[icao] = gdp2021
        gdp_2024_dict[icao] = gdp2024

    row += 1

# print("Population:")
# print(population_2021_dict)

# print("GDP 2021:")
# print(gdp_2021_dict)

# print("Population 2024:")
# print(population_2024_dict)

# print("GDP 2024:")
# print(gdp_2024_dict)


wb = load_workbook("DemandGroup40.xlsx", data_only=True)
ws = wb.active 

icao_row = 5    
lat_row = 6     
lon_row = 7    
start_col = 3   

airports = []
latitudes = []
longitudes = []

col = start_col
while True:
    icao = ws.cell(row=icao_row, column=col).value
    lat = ws.cell(row=lat_row, column=col).value
    lon = ws.cell(row=lon_row, column=col).value
    if icao is None:
        break
    airports.append(icao)
    latitudes.append(float(lat))
    longitudes.append(float(lon))
    col += 1


RE = 6371.0 

def distance(phi_i, lam_i, phi_j, lam_j):
    phi_i, phi_j = np.radians(phi_i), np.radians(phi_j)
    lam_i, lam_j = np.radians(lam_i), np.radians(lam_j)
    return 2 * RE * np.arcsin(
        np.sqrt(np.sin((phi_i - phi_j)/2)**2 +np.cos(phi_i)*np.cos(phi_j)*np.sin((lam_i - lam_j)/2)**2))


n = len(airports)
dij = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        dij[i, j] = distance(latitudes[i], longitudes[i], latitudes[j], longitudes[j])


# print("Afstandsmatrix (km):")
# print("\t" + "\t".join(airports))
# for i, row in enumerate(dij):
#     print(airports[i], "\t" + "\t".join(f"{val:.2f}" for val in row))


demand_start_row = icao_row + 8 
demand_start_col = start_col 

n = len(airports)
D = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        cell = ws.cell(row=demand_start_row + i, column=demand_start_col + j).value
        try:
            D[i, j] = float(cell) if cell is not None else 0
        except:
            D[i, j] = 0

# print("Demandmatrix 2021 (week):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D[i,j]:.0f}" for j in range(n)))

# Determning B1, B2, B3 and k
f = 1.42   # fuel cost constant from assignment

n = len(airports)
rows = []

for i in range(n):        
    for j in range(n):
        if i == j:
            continue                    # If orgin and destionation is the same, go on

        Dij = D[i, j]
        if Dij <= 0:
            continue                    # It there is no demand, go on

        i_icao = airports[i]            #Collectiong ICAO code
        j_icao = airports[j]

        row = {
            "D": Dij,
            "pop": population_2021_dict[i_icao] * population_2021_dict[j_icao],
            "gdp": gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao],
            "fd": f * dij[i, j]
        }
        rows.append(row)

df = pd.DataFrame(rows)
# print(df)

# --- LOGS ---
df['lnD']    = np.log(df['D'])                  #logarithm to lineariz the model, dat doe je door ln van alles te nemen heeft google mij verteld
df['ln_pop'] = np.log(df['pop'])
df['ln_gdp'] = np.log(df['gdp'])
df['ln_fd']  = np.log(df['fd'])

# --- OLS REGRESSIE ---
X = df[['ln_pop', 'ln_gdp', 'ln_fd']]           # Independent variables
X = sm.add_constant(X)                          # This creates ln(k)
y = df['lnD']                                   # Dependent variable

model = sm.OLS(y, X).fit()                      # Ordinary least squares
# print(model.summary())

# --- PARAMETERS ---
a = model.params['const']            # ln(k)
b1 = model.params['ln_pop']          # Creating b1
b2 = model.params['ln_gdp']

beta3 = model.params['ln_fd']        # = -b3 (afstand in noemer)
b3 = -beta3                          # Because in function it is in the denominator

k = np.exp(a)                       # Making k again after it is a ln()

# print("\n--- GRAVITY MODEL PARAMETERS ---")
# print(f"k  = {k}")
# print(f"b1 = {b1:.4f}")
# print(f"b2 = {b2:.4f}")
# print(f"b3 = {b3:.4f}")

# Maak een lege matrix voor voorspelde demand
D_pred = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred[i, j] = 0  # geen vraag van een stad naar zichzelf
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product = population_2021_dict[i_icao] * population_2021_dict[j_icao]
        gdp_product = gdp_2021_dict[i_icao] * gdp_2021_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred[i, j] = k * (pop_product**b1) * (gdp_product**b2) * ((fd_ij)**(beta3))

# print("Voorspelde demand (gravity model):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D_pred[i,j]:.0f}" for j in range(n)))


# Verschilmatrix
D_diff = D_pred - D  # absoluut verschil

# print("Absoluut verschil (voorspeld - werkelijke vraag):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D_diff[i,j]:.0f}" for j in range(n)))

# POP EN GDP VOORSPELLEN 2026
population_2026_dict = {}
for icao in population_2021_dict:
    pop_2021 = population_2021_dict[icao]
    pop_2024 = population_2024_dict[icao]

    # jaarlijkse groeivoet
    g = (pop_2024 / pop_2021)**(1/3)  

    # voorspelling voor 2026
    pop_2026 = pop_2024 * (g)**2
    population_2026_dict[icao] = pop_2026

gdp_2026_dict = {}
for icao in gdp_2021_dict:
    gdp_2021 = gdp_2021_dict[icao]
    gdp_2024 = gdp_2024_dict[icao]

    # Grow factor, in 3 years de difference is grow
    g = (gdp_2024 / gdp_2021)**(1/3)

    # 2026 is amount of 2024 times the grow factor to the power of the difference in years
    gdp_2026 = gdp_2024 * (g)**2
    gdp_2026_dict[icao] = gdp_2026

# print(f'Population in 2026')
# print(gdp_2026_dict)

# print(f'GDP in 2026')
# print(population_2026_dict)

# Demand prediction 2026
D_pred_26 = np.zeros((n, n))

for i in range(n):
    for j in range(n):
        if i == j:
            D_pred_26[i, j] = 0  # geen vraag van een stad naar zichzelf
            continue
        i_icao = airports[i]
        j_icao = airports[j]

        pop_product_26 = population_2026_dict[i_icao] * population_2026_dict[j_icao]
        gdp_product_26 = gdp_2026_dict[i_icao] * gdp_2026_dict[j_icao]
        fd_ij = f * dij[i, j]

        D_pred_26[i, j] = k * (pop_product_26**b1) * (gdp_product_26**b2) * ((fd_ij)**(beta3))

# print("Voorspelde demand 2026 (gravity model):")
# print("\t" + "\t".join(airports))
# for i in range(n):
#     print(airports[i], "\t" + "\t".join(f"{D_pred_26[i,j]:.0f}" for j in range(n)))

# Voorspelde ln(D) met het model
df['lnD_pred'] = model.predict(X)

# Omzetten naar originele schaal
df['D_pred'] = np.exp(df['lnD_pred'])
df['D_actual'] = np.exp(df['lnD'])

# # Plot: Werkelijk vs Voorspeld D
# plt.figure(figsize=(8,6))
# plt.scatter(df['D_actual'], df['D_pred'], color='blue', alpha=0.6, label='Data points')
# plt.plot([0, df['D_actual'].max()], [0, df['D_actual'].max()],
#          color='red', lw=2, label='Regression line')
# plt.xlabel('Given demand')
# plt.ylabel('Predicted demand')
# plt.title('Given vs. predicted demand 2021')
# plt.legend()
# plt.grid(True)
# plt.show()


### Implementatie van mathematical model ###

from gurobipy import *

# Inladen aircraft
# --- Workbook openen ---
wb2 = openpyxl.load_workbook("AircraftData.xlsx")
sheet2 = wb2.active   # neem het eerste werkblad

# Lees de header (aircraft names)
aircraft_names = [cell.value for cell in sheet2[1][1:]]  # rij 1, vanaf kolom B

data = {}

# Loop door rijen en vul dictionary
for row in sheet2.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        continue
    
    parameter_name = row[0]
    values = row[1:]
    
    data[parameter_name] = values

# DataFrame bouwen (transpose zodat aircraft types rijen worden)
df_aircraft = pd.DataFrame(data, index=aircraft_names)
df_aircraft = df_aircraft.drop(columns=["Aircraft Characteristics"])


# Data - Sets
N = range(len(airports))                    # Set of airports; i, j in N
K = range(len(df_aircraft))                 # Set of aircrafts; k in K
ac = len(df_aircraft)                        # Total amount of aircrafts

# Data - parameters
q = np.zeros((n, n))                        # Demand between orgin i and destination j
for i in N:
    for j in N:
        q[i, j] = D_pred_26[i, j]

d = np.zeros((n, n))                        # Distance between i and j
for i in N:
    for j in N:
        d[i, j] = dij[i, j]

y = np.zeros((n, n))                        # Yield in euro betwee i and j
for i in N:
    for j in N:
        y[i, j] = 5.9 * d[i, j]**(-0.76) + 0.043

s = np.zeros(ac)
for k in K:
    s[k] = df_aircraft['Seats'][k]

v = np.zeros(ac)
for k in K:
    v[k] = df_aircraft['Speed [km/h]'][k]

t = np.zeros(ac)
for k in K:
    t[k] = df_aircraft['Average TAT [mins]'][k]

ra = np.zeros(ac)
for k in K:
    ra[k] = df_aircraft['Maximum range [km]'][k]

RAC = np.zeros(ac)
for k in K:
    RAC[k] = df_aircraft['Runway required [m]'][k]

cl = np.zeros(ac)
for k in K:
    cl[k] = df_aircraft['Weekly lease cost [€]'][k]

C = np.zeros(ac)
for k in K:
    C[k] = df_aircraft['Fixed operating cost C_X [€]'][k]

CT = np.zeros(ac)
for k in K:
    CT[k] = df_aircraft['Time cost parameter C_T [€/hr]'][k]

C_Tij = np.zeros((ac, n, n))
for k in K:
    C_Tij[k, :, :] = CT[k] * (d / v[k])

CF = np.zeros(ac)
for k in K:
    CF[k] = df_aircraft['Fuel cost parameter C_F'][k]

C_Fij = np.zeros((ac, n, n))
for k in K:
    C_Fij[k, :, :] = ((CF[k] * 1.42) / 1.5) * d

for k in K:
    print(f'For Aircraf {k} fuel cost is {C_Fij[k]}')
