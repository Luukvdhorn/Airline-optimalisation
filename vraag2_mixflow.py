from gurobipy import * 
from openpyxl import * 
import openpyxl
from time import *
import numpy as np
import math
import matplotlib.pyplot as plt
import pandas as pd
import statsmodels.api as sm

import pandas as pd
import numpy as np
from openpyxl import load_workbook


# DATA INLADEN
# wb_data = load_workbook("data_ex2.xlsx", data_only=True)
wb_data = load_workbook("Group_2.xlsx", data_only=True)

flights_sheet = wb_data["Flights"]
itineraries_sheet = wb_data["Itineraries"]
recapture_sheet = wb_data["Recapture"]

# Zet sheets om naar DataFrames
df_flights = pd.DataFrame(flights_sheet.values)
df_flights.columns = df_flights.iloc[0]
df_flights = df_flights.iloc[1:].reset_index(drop=True)

df_itineraries = pd.DataFrame(itineraries_sheet.values)
df_itineraries.columns = df_itineraries.iloc[0]
df_itineraries = df_itineraries.iloc[1:].reset_index(drop=True)

df_recapture = pd.DataFrame(recapture_sheet.values)
df_recapture.columns = df_recapture.iloc[0]
df_recapture = df_recapture.iloc[1:].reset_index(drop=True)


# SETS
L = range(len(df_flights))        # flights
P = range(len(df_itineraries))    # itineraries

# Pp: mogelijke recapture-itineraries per p
Pp = {p: set() for p in P}
for _, row in df_recapture.iterrows():
    p = int(row["From Itinerary"])
    r = int(row["To Itinerary"])
    Pp[p].add(r)


# PARAMETERS

# fare_p
fare = np.zeros(len(P))
for p in P:
    fare[p] = float(df_itineraries.loc[p, "Price [EUR]"])

# Demand_p
D = np.zeros(len(P))
for p in P:
    D[p] = float(df_itineraries.loc[p, "Demand"])

# CAP_l
CAP = np.zeros(len(L))
for l in L:
    CAP[l] = float(df_flights.loc[l, "Capacity"])


# b_p^r  (recapture rates)
b_pr = {}
for _, row in df_recapture.iterrows():
    p = int(row["From Itinerary"])
    r = int(row["To Itinerary"])
    b_pr[(p, r)] = float(row["Recapture Rate"])


# Maak mapping van flight code naar index l
flight_code_to_id = {df_flights.loc[l, 'Flight No.']: l for l in L}

# delta_lp = 1 als flight l in itinerary p zit
delta_lp = {}

for p, row in df_itineraries.iterrows():

    # Flight 1 (altijd aanwezig)
    code1 = row['Flight 1']
    if pd.notna(code1):
        l = flight_code_to_id[code1]
        delta_lp[(l, p)] = 1

    # Flight 2 (optioneel)
    code2 = row['Flight 2']
    if pd.notna(code2):
        l = flight_code_to_id[code2]
        delta_lp[(l, p)] = 1



# CONTROLE

from gurobipy import Model, GRB, quicksum

def main():
    model = Model("Passenger_Mix_Flow")
    model.setParam("TimeLimit", 300)

    # DECISION VARIABLES
    x = model.addVars(P, P, lb=0, vtype=GRB.INTEGER, name="x")

    # OBJECTIVE FUNCTION
    model.setObjective(quicksum(fare[r] * x[r, p] for p in P for r in P), GRB.MAXIMIZE)

    # CAPACITY CONSTRAINTS
    for l in L:
        model.addConstr(quicksum(delta_lp.get((l, r), 0) * x[r, p] for p in P for r in P) <= CAP[l],
            name=f"cap_{l}")

    # DEMAND CONSTRAINTS
    for p in P:
        model.addConstr(quicksum(x[r, p] / b_pr[(p, r)]for r in Pp[p]) <= D[p],
            name=f"demand_{p}")

    # OPTIMIZE
    model.optimize()

    # OUTPUT
    if model.status == GRB.OPTIMAL or model.status == GRB.TIME_LIMIT:
        print(f"Objective value: {model.objVal}")
        for r, p in x.keys():
            if x[r, p].x > 1e-6:
                print(f"x[{r},{p}] = {x[r,p].x}")
    else:
        print("No optimal solution found.")

main()

# def main():
#     model = Model("Model_mix_flow")
#     model.setParam('TimeLimit', 300)

# #Decision variables
#     x = {}
#     for p in P:
#         for r in Pp[p] | {p}:  # p kan altijd "recapturen" naar zichzelf
#             x[r, p] = model.addVar(vtype=GRB.INTEGER, lb=0, name=f"x_{r}_{p}")
#     model.update()



# #Objective function
#     objective_2_maxflow = quicksum((fare[p] * x[r, p] ) for p in P for r in Pp[p] | {p})

#     model.setObjective(objective_2_maxflow, GRB.MAXIMIZE)

# #Constraints
#     #capacity constraint
#     for l in L:
#         model.addConstr(quicksum(x[r, p] * delta_lp.get((l, p), 0) for p in P for r in P) <= CAP[l], 
#                         name=f"cap_{l}")
#     #demand constraint
#     for p in P:
#         model.addConstr(quicksum(x[r, p] / b_pr.get((p,r), 1) for r in Pp[p] | {p}) <= D[p],
#                         name=f"demand_{p}")
#     # #Nonzero constraint
#     # for p in P:
#     #     for r in Pp[p] | {p}:
#     #         model.addConstr(x[r, p] >= 0, name=f"nonneg_{r}_{p}")

# #optimize model
#     model.optimize()
#     if model.status == GRB.OPTIMAL or model.status == GRB.TIME_LIMIT:
#         print(f"Optimal objective value: {model.objVal}")
#         for v in model.getVars():
#             if v.x > 0:
#                 print(f"{v.varName}: {v.x}")

#     else:
#         print("No optimal solution found.")


# main()